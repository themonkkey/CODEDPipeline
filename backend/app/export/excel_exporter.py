"""
Build a single Excel workbook from extracted tables:
one sheet per table (heading row = title from the PDF),
plus an Index sheet with hyperlinks to every table.
"""

import io
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="C8102E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
LINK_FONT = Font(color="0563C1", underline="single")


# words too generic to make a sheet tab recognisable on their own
_STOPWORDS = {
    "table", "tabel", "statement", "annexure", "appendix", "of", "the", "and",
    "in", "on", "for", "to", "by", "at", "a", "an", "no", "per", "number",
    "percentage", "distribution",
}


def _slug_words(title):
    """A few distinctive words from a title, for a human-readable sheet tab."""
    words = re.findall(r"[A-Za-z]{3,}", str(title))
    keep = [w for w in words if w.lower() not in _STOPWORDS]
    return keep or words


def _sheet_name(title, table_id, used):
    """Excel sheet names: <=31 chars, unique, no \\/?*[]: characters.

    Build a navigable tab from the title — a "Table X.Y" code plus a couple of
    distinctive words ("T6_2_Sex_Ratio") — so a 600-sheet workbook can be skimmed
    by tab alone. Falls back to the table id only when there is no usable title.
    """
    title = str(title or "")

    m = re.search(r"(?:Table|Tabel|Statement)\s+(\d+(?:[.\-]\d+)*)", title, re.IGNORECASE)
    slug = "_".join(_slug_words(title)[:3])

    if m:
        code = "T" + m.group(1).replace(".", "_").replace("-", "_")
        base = f"{code}_{slug}" if slug else code
    elif slug:
        base = slug
    else:
        base = f"table_{table_id}"

    base = re.sub(r"[\\/?*\[\]:]", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")[:28] or f"table_{table_id}"

    name, n = base, 2
    while name in used:
        suffix = f"_{n}"
        name = base[:28 - len(suffix)] + suffix
        n += 1

    used.add(name)
    return name


def _maybe_number(value):
    """Store numerics as numbers so Excel can compute on them."""

    text = str(value).strip().replace(",", "")
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d*\.\d+", text):
        return float(text)
    return value


def build_workbook(table_dfs, catalog):
    """
    table_dfs: {table_id: DataFrame}
    catalog:   list of metadata dicts with table_id, table_name, page
    Returns BytesIO of the .xlsx file.
    """

    meta = {m["table_id"]: m for m in catalog}

    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "Contents"

    used_names = set()
    index_rows = []

    for tid, df in table_dfs.items():

        # disk-backed mode: value may be a CSV path — load one at a time
        if isinstance(df, str):
            df = pd.read_csv(df, dtype=str, keep_default_na=False)

        # never emit a tab for an empty/ghost frame
        if df is None or df.shape[0] == 0 or df.shape[1] == 0:
            continue

        info = meta.get(tid, {})
        title = info.get("table_name") or f"table_{tid}"
        page = info.get("page", "")

        sheet = _sheet_name(title, tid, used_names)
        ws = wb.create_sheet(sheet)

        n_cols = max(df.shape[1], 1)

        # title row (from the PDF), merged across the table width
        ws.cell(row=1, column=1, value=str(title)).font = TITLE_FONT
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=n_cols
        )

        # column header row
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=c, value=str(col))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # data
        for r, row in enumerate(df.itertuples(index=False), start=4):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=_maybe_number(value))

        # column widths
        for c, col in enumerate(df.columns, start=1):
            width = max(
                len(str(col)),
                *(len(str(v)) for v in df.iloc[:, c - 1].head(50)),
                8,
            )
            ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 45)

        ws.freeze_panes = "A4"
        index_rows.append((tid, title, page, sheet, df.shape[0], df.shape[1]))

    # ---- Contents / TOC sheet (first tab): full untruncated titles, one
    #      clickable row per table, mapping the short sheet code back to the
    #      real title plus page and dimensions so a big workbook stays navigable.
    headers = ["#", "Table Name", "PDF Page", "Rows", "Cols", "Sheet"]
    for c, h in enumerate(headers, 1):
        cell = index_ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for r, (tid, title, page, sheet, nrows, ncols) in enumerate(index_rows, start=2):
        index_ws.cell(row=r, column=1, value=r - 1)          # running index
        link = index_ws.cell(row=r, column=2, value=str(title))
        link.hyperlink = f"#'{sheet}'!A1"
        link.font = LINK_FONT
        index_ws.cell(row=r, column=3, value=page)
        index_ws.cell(row=r, column=4, value=nrows)
        index_ws.cell(row=r, column=5, value=ncols)
        index_ws.cell(row=r, column=6, value=sheet)

    widths = {"A": 6, "B": 80, "C": 9, "D": 7, "E": 6, "F": 30}
    for col, w in widths.items():
        index_ws.column_dimensions[col].width = w
    index_ws.freeze_panes = "A2"

    # guarantee a visible sheet even when nothing was exported (openpyxl refuses
    # to save a workbook with no visible sheet)
    if not index_rows:
        index_ws.cell(row=2, column=2, value="No tables extracted.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
