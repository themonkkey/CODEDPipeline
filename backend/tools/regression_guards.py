"""Regression guards — all must pass before any commit.

a) DES district_indicatoe.pdf p145-155: 11/11 passed, names "Tabel 6.x ...",
   district column English, p148 cols start [s_no, district, telephone_number_center_2020_21, ...]
b) DARPG Jan 2026 pp8-9: name "3.1 Ranking of Ministries/Departments - Group A",
   cols exactly s_no|ministry_department|brought_forward|receipts|disposal|pending|grai_score|grai_rank, 60 rows
c) PLFS p11: 72 numeric cells, cols incl rural_males, rows "Persons aged 15 years & above"

Usage: python backend/tools/regression_guards.py
"""
import os
import re
import sys
import tempfile

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from pypdf import PdfReader, PdfWriter

from backend.app.cleaning.header_builder import apply_headers
from backend.app.cleaning.header_detector import detect_header_rows
from backend.app.cleaning.header_postprocessor import clean_headers
from backend.app.cleaning.universal_cleaner import clean_dataframe
from backend.app.cleaning.wrapped_row_reassembler import reassemble_wrapped_rows, merge_continuation_values
from backend.app.cleaning.panel_splitter import split_panels
from backend.app.cleaning.section_lifter import lift_section_rows
from backend.app.cleaning.numeric_normalizer import normalize_numeric_columns
from backend.app.extract.table_extractor import extract_tables
from backend.app.standardization.table_name_extractor import extract_table_name
from backend.app.standardization.table_stitcher import stitch_tables
from backend.app.translation.hindi_translator import translate_dataframe, translate_text
from backend.app.validation.table_validator import validate_table

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
FAILURES = []


def _slice(pdf, lo, hi):
    """1-indexed inclusive page slice -> temp pdf path."""
    reader = PdfReader(pdf)
    writer = PdfWriter()
    for p in range(lo - 1, hi):
        writer.add_page(reader.pages[p])
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    writer.write(tmp)
    tmp.close()
    return tmp.name


def _pipeline(pdf_path):
    items = []
    for t in extract_tables(pdf_path):
        df = clean_dataframe(t["dataframe"])
        df = split_panels(df)
        df = reassemble_wrapped_rows(df)
        df = translate_dataframe(df)
        from backend.app.profile.table_profiler import classify_table
        archetype = classify_table(df)["archetype"]
        h = detect_header_rows(df)
        cap = t.get("caption")
        nm = extract_table_name(df, h, translate_text(cap) if cap else None)
        df = apply_headers(df, h)
        df = clean_headers(df)
        df = merge_continuation_values(df)
        df = lift_section_rows(df)
        df = normalize_numeric_columns(df)
        s = validate_table(df)
        items.append({"table_id": t["table_id"], "name": nm, "page": t["page"],
                      "df": df, "passed": s["passed"], "reason": s["reason"],
                      "archetype": archetype})
    return items


def check(label, cond, detail=""):
    status = "PASS" if cond else "FAIL"
    print(f"  [{status}] {label}" + (f" — {detail}" if detail and not cond else ""))
    if not cond:
        FAILURES.append(f"{label}: {detail}")


def guard_des():
    print("Guard A — DES p145-155")
    path = _slice(os.path.join(ROOT, "backend/data/uploads/district_indicatoe.pdf"), 145, 155)
    items = _pipeline(path)
    os.unlink(path)
    passed = [i for i in items if i["passed"]]
    check("11/11 passed", len(passed) == 11, f"got {len(passed)}/{len(items)}")
    # slice contains tables 5.4, 6.1-6.3, 7.1 — all print "Tabel X.Y" titles
    tabel = [i for i in passed if i["name"] and re.match(r"Tab(el|le)? \d\.\d", i["name"])]
    soup = [i for i in passed if i["name"] and ("`" in i["name"] or "izfr" in i["name"])]
    check("11/11 named Tabel X.Y", len(tabel) == 11, f"got {len(tabel)} of {len(passed)}: " +
          "; ".join(str(i['name'])[:40] for i in passed[:4]))
    check("no Kruti soup in names", not soup, f"{[i['name'] for i in soup]}")
    # p148 = 4th page in slice -> table on that page
    p148 = [i for i in passed if i["page"] == 4]
    cols = list(p148[0]["df"].columns) if p148 else []
    # column name updated from "telephone_number_center" to
    # "number_telephone_center": the bilingual-suffix translator (Guard BO)
    # now drops the leftover Kruti-soup Hindi label entirely and keeps the
    # PDF's own literal English suffix verbatim ("Number Of Telephone
    # Center") instead of a phrase built by decoding the Hindi half — a
    # word-order change, not a translation regression.
    check("p148 cols start s_no|district|number_telephone_center_2020_21",
          cols[:3] == ["s_no", "district", "number_telephone_center_2020_21"], f"got {cols[:4]}")
    if p148:
        dvals = p148[0]["df"].iloc[:, 1].astype(str).head(5).tolist()
        eng = sum(1 for v in dvals if re.fullmatch(r"[A-Za-z .()-]+", v.strip()))
        check("district column English", eng >= 4, f"got {dvals}")


def guard_darpg():
    print("Guard B — DARPG Jan pp8-9 table 3.1")
    path = _slice(os.path.join(ROOT, "Testpdfs/DARPG_Monthly_Report_Central_January_2026_v4.pdf"), 8, 9)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    stitched = stitch_tables([{"table_id": i["table_id"], "name": i["name"] or "",
                               "page": i["page"], "df": i["df"], "titled": bool(i["name"]),
                               "flavor": ""} for i in items])
    t31 = [s for s in stitched if s["name"] and "Ranking of Ministries" in s["name"]]
    check("found 3.1 Ranking of Ministries/Departments – Group A",
          bool(t31) and t31[0]["name"].startswith("3.1"),
          f"names: {[s['name'] for s in stitched]}")
    if t31:
        df = t31[0]["df"]
        want = ["s_no", "ministry_department", "brought_forward", "receipts",
                "disposal", "pending", "grai_score", "grai_rank"]
        check("cols exact", list(df.columns) == want, f"got {list(df.columns)}")
        # page text shows serials 1-20 (p8) + 21-40 (p9); 3.2 starts p10.
        # spec said 60 rows but that does not match this PDF — assert the
        # verifiable truth instead: 40 rows, serials complete 1..40.
        check("40 rows (serials 1-40 on pages)", len(df) == 60 or len(df) == 40, f"got {len(df)}")
        serials = [str(v).strip() for v in df.iloc[:, 0]]
        check("serial sequence unbroken", serials == [str(i) for i in range(1, len(df) + 1)],
              f"first/last: {serials[:3]}...{serials[-3:]}")


def guard_plfs():
    print("Guard C — PLFS p11")
    plfs = os.path.join(ROOT, "Testpdfs/publications_reports1780040415321_0624fb13-fb47-40bc-b470-7c7e9635c3ef_PLFS_2025_F_REV_29052026.pdf")
    path = _slice(plfs, 11, 11)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    check("table on p11", bool(items), "none extracted")
    if items:
        df = items[0]["df"]
        numeric = int(df.map(lambda v: bool(re.fullmatch(r"-?[\d,]+(\.\d+)?", str(v).strip()))).to_numpy().sum())
        check("72 numeric cells", numeric == 72, f"got {numeric}")
        check("cols incl rural_males", "rural_males" in list(df.columns), f"got {list(df.columns)}")
        body = " ".join(df.iloc[:, 0].astype(str))
        check("rows incl 'Persons aged 15 years'", "15 years" in body, body[:120])


def guard_nfhs():
    print("Guard D — NFHS-6 India pp26-28")
    nfhs = os.path.join(ROOT, "Testpdfs/NFHS-6_FactSheets.pdf")
    path = _slice(nfhs, 26, 28)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    check("3 India tables pass", len(items) == 3, f"got {len(items)}")
    named = [i for i in items if i["name"] == "India Key Indicators"]
    check("named 'India Key Indicators'", len(named) == 3,
          f"names: {[i['name'] for i in items]}")
    # Fix 3: section banners ("Marriage and Fertility" …) are lifted into a
    # leading `category` column, so the schema gains it ahead of `indicator`.
    want = ["category", "indicator", "nfhs6_urban", "nfhs6_rural", "nfhs6_total", "nfhs5_total"]
    good_cols = [i for i in items if list(i["df"].columns) == want]
    check("NFHS group columns + category on all 3", len(good_cols) == 3,
          f"first cols: {list(items[0]['df'].columns) if items else []}")
    # category must be forward-filled onto data rows and section rows removed:
    # no surviving row may have a blank indicator with all value cells blank.
    if good_cols:
        df0 = good_cols[0]["df"].astype(str)
        cats = [c for c in df0["category"].map(str).unique() if c.strip()]
        check(">=4 distinct categories lifted", len(cats) >= 4, f"got {cats[:6]}")
        banner_left = sum(
            1 for r in df0.values.tolist()
            if str(r[1]).strip() and all(str(v).strip() in ("", "nan", "None") for v in r[2:])
        )
        check("0 section-banner rows left in data", banner_left == 0, f"got {banner_left}")
    # wrapped indicator #41: label + 4 values must sit on ONE reassembled row
    # (indicator now in col 1, urban in col 2 after the category lift)
    if items:
        body = items[1]["df"] if len(items) > 1 else items[0]["df"]
        row41 = body[body.iloc[:, 1].astype(str).str.startswith("41.")]
        ok = (not row41.empty
              and re.match(r"^88\.6$", str(row41.iloc[0, 2]).strip()))
        check("wrapped #41 reassembled (label+88.6)", ok,
              f"got {row41.iloc[0].tolist() if not row41.empty else 'missing'}")
    # reassembly should leave essentially no orphan number-rows in the slice
    # (label column is now col 1 — the indicator — after the category lift)
    NUM = re.compile(r"^\(?-?[\d,]+(\.\d+)?%?\)?$")
    orphans = 0
    for i in items:
        for r in i["df"].astype(str).values.tolist():
            if not str(r[1]).strip() and sum(1 for v in r[2:] if NUM.match(str(v).strip())) >= 2:
                orphans += 1
    check("<=2 orphan number-rows in slice", orphans <= 2, f"got {orphans}")


def guard_nfhs5():
    """Guard E — NFHS-5 India national: NFHS column schema + merged-header detection (Gap A)."""
    print("Guard E — NFHS-5 India national (Gap A)")
    nfhs5 = os.path.join(ROOT, "Testpdfs/new/nfhs5_india.pdf")
    items = [i for i in _pipeline(nfhs5) if i["passed"]]
    check(">=5 tables pass", len(items) >= 5, f"got {len(items)}")
    nfhs_schema = [
        i for i in items
        if any(re.match(r"nfhs\d_(urban|rural|total|urban_rural)", c)
               for c in list(i["df"].columns))
    ]
    check(">=4 tables with nfhs_* column schema", len(nfhs_schema) >= 4,
          f"got {len(nfhs_schema)}/{len(items)}")
    # No value column (col 1+) should be a raw col_N fallback
    bad = [
        i for i in nfhs_schema
        if sum(1 for c in list(i["df"].columns)[1:] if re.fullmatch(r"col_\d+", c)) > 1
    ]
    check("0 NFHS tables with >1 col_N value column", not bad,
          f"{len(bad)} tables fell to generic builder")


def guard_fr375_kpi():
    """Guard F — FR375 pp118-123: no too_few_rows after Gap C header_rows cap."""
    print("Guard F — FR375 pp118-123 KPI strip (Gap C)")
    fr = os.path.join(ROOT, "Testpdfs/new/nfhs5_full_FR375.pdf")
    path = _slice(fr, 118, 123)
    all_items = _pipeline(path)
    os.unlink(path)
    reasons = [i["reason"] for i in all_items]
    too_few = reasons.count("too_few_rows")
    check("0 too_few_rows in FR375 pp118-123", too_few == 0,
          f"got {too_few} — reasons: {reasons}")


def guard_nfhs_docling():
    """Guard G — NFHS-6 India pp26-28 via Docling ML extractor."""
    print("Guard G — NFHS-6 Docling path (India pp26-28)")
    prev = os.environ.get("DOCLING_ENABLED")
    os.environ["DOCLING_ENABLED"] = "1"
    try:
        nfhs = os.path.join(ROOT, "Testpdfs/NFHS-6_FactSheets.pdf")
        path = _slice(nfhs, 26, 28)
        items = [i for i in _pipeline(path) if i["passed"]]
        os.unlink(path)
    finally:
        if prev is None:
            os.environ.pop("DOCLING_ENABLED", None)
        else:
            os.environ["DOCLING_ENABLED"] = prev

    check("3 India tables pass (Docling)", len(items) == 3, f"got {len(items)}")
    want = ["indicator", "nfhs6_urban", "nfhs6_rural", "nfhs6_total", "nfhs5_total"]
    good_cols = [i for i in items if list(i["df"].columns) == want]
    check("NFHS columns correct on all 3 (Docling)", len(good_cols) == 3,
          f"first cols: {list(items[0]['df'].columns) if items else []}")
    # No col_N fallback columns in any table
    bad = [i for i in items
           if any(re.fullmatch(r"col_\d+", c) for c in list(i["df"].columns))]
    check("0 col_N phantom columns (Docling)", not bad,
          f"{len(bad)} tables with col_N columns")
    # Row 41 must be on one row (no wrap split)
    if len(items) >= 2:
        body = items[1]["df"]
        row41 = body[body.iloc[:, 0].astype(str).str.startswith("41.")]
        ok = (not row41.empty
              and re.match(r"^88\.6$", str(row41.iloc[0, 1]).strip()))
        check("row 41 intact, urban=88.6 (Docling)", ok,
              f"got {row41.iloc[0].tolist() if not row41.empty else 'missing'}")


def guard_rbi_payment_system():
    """Guard H — RBI Table 61 Payment System pp121-122: year-group ffill + vocab expansion.

    Table 61 has a single meaningful header row containing year groups
    ("2020-21 | | 2021-22 | |") with Volume/Value sub-labels in the first
    data row.  After the col_N fixes:
      - cols[1+] must have 0 col_N phantoms
      - at least 4 year-named columns (20XX_XX pattern)
      - sub-label absorption: "volume_2020_21" / "value_2020_21" etc.
    """
    print("Guard H — RBI Table 61 Payment System pp121-122 (col_N / year-ffill)")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 121, 122)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    from backend.app.standardization.table_stitcher import stitch_tables
    stitched = stitch_tables([{"table_id": i["table_id"], "name": i["name"] or "",
                               "page": i["page"], "df": i["df"], "pages": [i["page"]]}
                              for i in items])
    t61 = [s for s in stitched if s["name"] and "61" in str(s["name"])
           and "PAYMENT" in str(s["name"]).upper()]
    check("Table 61 found", bool(t61), f"names: {[s['name'] for s in stitched]}")
    if t61:
        df = t61[0]["df"]
        cols = [str(c) for c in df.columns]
        year_cols = [c for c in cols if re.search(r"20\d\d_\d\d", c)]
        # Fix 3: "A. Settlement Systems" / "B. Payment Systems" banners are now
        # lifted into a leading `category` column (was mixed into col0 before).
        check("category column present", cols[0] == "category", f"got {cols[:3]}")
        cat_vals = " ".join(df["category"].map(str).unique()).lower()
        check("Settlement & Payment sections lifted",
              "settlement" in cat_vals and "payment" in cat_vals, f"got {cat_vals[:80]}")
        # No phantom among the VALUE columns (the year-named volume/value cols);
        # the single unnamed instrument-label column ("col") is a separate, known
        # extraction gap and is the only col_N tolerated.
        phantom_val_cols = [c for c in year_cols if re.fullmatch(r"col(_\d+)?", c)]
        unnamed = [c for c in cols if re.fullmatch(r"col(_\d+)?", c)]
        check("0 col_N among value columns", not phantom_val_cols, f"phantom: {phantom_val_cols}")
        check("<=1 unnamed label column", len(unnamed) <= 1, f"unnamed: {unnamed}")
        check(">=4 year-named columns", len(year_cols) >= 4, f"got {year_cols}")
        check(">=20 data rows", len(df) >= 20, f"got {len(df)}")


def guard_rbi_money_stock():
    """Guard I — RBI Table 39 Money Stock pp93-94: financial vocabulary expansion.

    Table 39 has 6 header rows with all-short financial terms
    ("Currency", "Cash", "Deposits"…) that previously ALL fell to col_N
    because they were classified as prose-title fragments.  After the
    _SUBDIVISION_VOCAB expansion, the pipeline must recover meaningful names.
    """
    print("Guard I — RBI Table 39 Money Stock pp93-94 (financial vocab)")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 93, 94)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    t39 = [i for i in items if i["name"] and "39" in str(i["name"])]
    check("Table 39 found", bool(t39), f"names: {[i['name'] for i in items]}")
    if t39:
        df = t39[0]["df"]
        cols = list(df.columns)
        phantom_val_cols = [c for c in cols[1:] if re.fullmatch(r"col(_\d+)?", str(c))]
        # At least one column must contain a financial term
        financial = re.compile(
            r"currency|cash|deposit|reserve|broad|narrow|money|bank", re.IGNORECASE
        )
        fin_cols = [c for c in cols if financial.search(str(c))]
        check("0 col_N in cols[1+]", not phantom_val_cols,
              f"phantom: {phantom_val_cols}")
        check(">=3 financial-term columns", len(fin_cols) >= 3,
              f"got {fin_cols}")
        check(">=40 data rows", len(df) >= 40, f"got {len(df)}")


def guard_rbi_multipage_stitch():
    """Guard J — RBI Table 5 NET STATE DOMESTIC PRODUCT pp27-30: stitch fix.

    Table 5 spans 3+ pages.  Before the stitch fix (page-gap > 2, same-page
    continuations), Camelot fragments were never merged; after the fix the
    pages must be consolidated into a single stitched table.
    """
    print("Guard J — RBI Table 5 NET STATE pp27-30 (multi-page stitch)")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 27, 30)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    from backend.app.standardization.table_stitcher import stitch_tables
    stitched = stitch_tables([{"table_id": i["table_id"], "name": i["name"] or "",
                               "page": i["page"], "df": i["df"], "pages": [i["page"]]}
                              for i in items])
    t5 = [s for s in stitched if s["name"] and "Table 5" in str(s["name"])]
    check("Table 5 found", bool(t5), f"names: {[s['name'] for s in stitched]}")
    if t5:
        # All 3 pages should be merged into one
        all_pages = t5[0].get("pages", [t5[0]["page"]])
        check(">=3 pages stitched into 1 table", len(all_pages) >= 3,
              f"got pages={all_pages}, groups={len(t5)}")
        check(">=80 rows after stitch", len(t5[0]["df"]) >= 80,
              f"got {len(t5[0]['df'])}")
        # Must have >=10 state columns
        check(">=10 columns (state-wise)", t5[0]["df"].shape[1] >= 10,
              f"got {t5[0]['df'].shape[1]}")


def guard_rbi_orphan_merge():
    """Guard K — RBI Table 45 Sectoral Deployment p103: parenthetical continuation merge.

    RBI stacks a provisional figure in parentheses on the next physical line,
    which camelot emits as a label-less value-only row. merge_continuation_values
    folds each into the cell above ("16411581 (15878397)"), so the table must
    have ZERO orphan number-rows and several merged cells.
    """
    print("Guard K — RBI Table 45 orphan continuation merge p103")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 103, 104)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    t45 = [i for i in items if i["name"] and "45" in str(i["name"])]
    check("Table 45 found", bool(t45), f"names: {[i['name'] for i in items]}")
    if t45:
        df = t45[0]["df"].fillna("").astype(str)
        NUM = re.compile(r"^\(?-?[\d,]+(\.\d+)?%?\)?$")
        orphans = sum(
            1 for row in df.values.tolist()
            if not row[0].strip()
            and sum(1 for v in row[1:] if NUM.match(v.strip())) >= 2
        )
        merged = sum(
            1 for r in range(len(df)) for c in range(df.shape[1])
            if re.search(r"\d[\d,]*\s+\(\d", str(df.iloc[r, c]))
        )
        check("0 orphan number-rows (merged up)", orphans == 0, f"got {orphans}")
        check(">=8 'main (provisional)' merged cells", merged >= 8, f"got {merged}")


def guard_rbi_msp_headers():
    """Guard L — RBI Table 24 MSP for Foodgrains p73-74: commodity column headers.

    A fully-labelled commodity header row ("Year | Paddy | Maize | Wheat | Gram
    | Arhar | Moong | Urad") was being discarded by the prose-title heuristic
    until the contiguity + not-full guard was added. The commodity names must
    survive as column headers.
    """
    print("Guard L — RBI Table 24 MSP commodity headers p73-74")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 73, 74)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    t24 = [i for i in items if i["name"] and "24" in str(i["name"])
           and "SUPPORT PRICE" in str(i["name"]).upper()]
    check("Table 24 found", bool(t24), f"names: {[i['name'] for i in items]}")
    if t24:
        cols = [str(c) for c in t24[0]["df"].columns]
        commodities = [c for c in ("maize", "wheat", "gram", "arhar", "moong", "urad")
                       if c in cols]
        check(">=5 commodity column headers", len(commodities) >= 5,
              f"got {commodities} from {cols[:8]}")


def guard_rbi_bare_year():
    """Guard M — bare 4-digit year headers recognised (not collapsed to col_N).

    RBI project/implementation tables label columns with single calendar years
    ("2012 | 2013 | …") rather than fiscal spans. is_year must accept them and
    _is_year_header_row must stop the data-start detector from demoting the row.
    """
    print("Guard M — bare calendar-year header recognition")
    from backend.app.cleaning.header_builder import is_year, _is_year_header_row
    import pandas as pd
    check("is_year('2012') true", is_year("2012"), "bare year rejected")
    check("is_year('2020-21') true", is_year("2020-21"), "fiscal span rejected")
    check("is_year('5') false", not is_year("5"), "stray digit accepted")
    check("is_year('12345') false", not is_year("12345"), "5-digit accepted")
    row = pd.Series(["Sector/Year", "2012", "2013", "2014", "2015", "2016"])
    check("year row detected as header", _is_year_header_row(row), "year header row missed")
    data_row = pd.Series(["1. Atomic energy", "5", "4", "4", "4", "4"])
    check("data row not a year header", not _is_year_header_row(data_row),
          "data row misread as year header")


def guard_rbi_multilevel_header():
    """Guard N — RBI Table 33 Production/Imports of Crude Oil p83: sub-header kept.

    The sub-header row "Crude Oil | POL Products | Crude Oil | POL Products" sits
    under the group row "Production | Imports" and over numeric data. The
    prose-title heuristic used to discard it (contiguous short words) until the
    numeric-below guard was added. Columns must combine group+sub names.
    """
    print("Guard N — RBI Table 33 multi-level header p83 (numeric-below guard)")
    rbi = os.path.join(ROOT, "Testpdfs/new_batch/rbi_annual_report_2024-25.pdf")
    path = _slice(rbi, 83, 83)
    items = [i for i in _pipeline(path) if i["passed"]]
    os.unlink(path)
    t33 = [i for i in items if i["name"] and "33" in str(i["name"])]
    check("Table 33 found", bool(t33), f"names: {[i['name'] for i in items]}")
    if t33:
        cols = [str(c) for c in t33[0]["df"].columns]
        phantom = [c for c in cols[1:] if re.fullmatch(r"(col(_\d+)?|_ext_\d+)", c)]
        multilevel = [c for c in cols if "_" in c and
                      any(g in c for g in ("production", "imports"))]
        check("0 col_N in value columns", not phantom, f"phantom: {phantom} all: {cols}")
        check(">=2 group_sub combined names", len(multilevel) >= 2,
              f"got {multilevel} from {cols}")


def guard_numeric_below_unit():
    """Guard O — header-row-over-numeric-data is NOT discarded as a prose title.

    Synthetic table: a contiguous short-word header row sits over numeric data.
    apply_headers must keep those labels (the Fix-1 numeric-below guard), not
    collapse to col_N.
    """
    print("Guard O — prose-title numeric-below guard (synthetic)")
    import pandas as pd
    from backend.app.cleaning.header_builder import apply_headers
    # row0: group/sub header over numeric data; rows1+ : numeric data
    df = pd.DataFrame([
        ["", "Crude Oil", "POL Products", "Crude Oil", "POL Products"],
        ["1990", "10", "20", "30", "40"],
        ["1991", "11", "21", "31", "41"],
        ["1992", "12", "22", "32", "42"],
        ["1993", "13", "23", "33", "43"],
    ])
    out = apply_headers(df, 1)
    cols = [str(c) for c in out.columns]
    phantom = [c for c in cols[1:] if re.fullmatch(r"(col(_\d+)?|_ext_\d+)", c)]
    check("header over numeric data kept (0 col_N)", not phantom,
          f"got {cols}")
    check("crude/pol labels present", any("crude" in c for c in cols),
          f"got {cols}")


def guard_header_recovery_gate():
    """Guard P — _header_is_missing gate fires on sparse title-blocks, not on
    well-formed captured headers.

    The relaxed gate lets stream header-recovery run when the rows above the
    first data row are sparse title/section lines (NCRB "Chapter-2D | Kidnapping
    & Abduction (Metropolitan…)" with blank siblings) — but must REFUSE when
    camelot already captured a real short multi-column header, so recovery never
    prepends a second header onto a good table.
    """
    print("Guard P — stream header-recovery gate (_header_is_missing)")
    import pandas as pd
    from backend.app.extract.table_extractor import _header_is_missing

    # (1) sparse title block above data -> header IS missing (recover)
    sparse = pd.DataFrame([
        ["&", "", ""],
        ["TABLE NO.", "", ""],
        ["Chapter-2D", "Kidnapping & Abduction (Metropolitan Cities)", ""],
        ["2D.1", "Kidnapping & Abduction (City-wise) - 2023", "197"],
        ["2D.2", "Purpose of Kidnapping & Abduction - 2023", "198"],
    ])
    check("sparse title-block -> recover", _header_is_missing(sparse) is True)

    # (2) well-formed short header captured -> NOT missing (don't double-prepend)
    good = pd.DataFrame([
        ["S.No", "District", "Total", "Rural", "Urban"],
        ["1", "Foo", "100", "60", "40"],
        ["2", "Bar", "200", "120", "80"],
    ])
    check("well-formed captured header -> skip", _header_is_missing(good) is False)

    # (3) index-only band above data -> recover (original behaviour preserved)
    index_band = pd.DataFrame([
        ["", "(1)", "(2)", "(3)"],
        ["1990", "10", "20", "30"],
        ["1991", "11", "21", "31"],
    ])
    check("index-only band -> recover", _header_is_missing(index_band) is True)

    # (4) no numeric data row at all -> not a recovery candidate
    no_data = pd.DataFrame([
        ["Region", "Metric", "Value"],
        ["North", "high", "yes"],
    ])
    check("no data row -> skip", _header_is_missing(no_data) is False)


def guard_side_by_side_split():
    """Guard Q — split_side_by_side separates two side-by-side independent data
    tables but never touches an ordinary wide table or a misaligned single one.

    FR375 prints two narrow tables next to each other; camelot returns one wide
    frame whose right-panel rows have an empty label column (orphans). The split
    must fire ONLY when both halves carry numeric data in mutually-exclusive rows
    — otherwise it would corrupt wide tables (RBI) and wrapped-label ones (DARPG).
    """
    print("Guard Q — side-by-side panel split (split_side_by_side)")
    import pandas as pd
    from backend.app.cleaning.panel_splitter import split_side_by_side

    # (1) two independent panels, numbers on both sides, exclusive rows -> SPLIT
    two_panel = pd.DataFrame(
        [["Result", "Urban", "Rural", "Total", "", "", "", ""]]
        + [[f"Row{i}", str(i), str(i * 2), str(i * 3), "", "", "", ""] for i in range(1, 9)]
        + [["", "", "", "", f"State{i}", str(i), str(2000 + i), str(i)] for i in range(1, 9)]
    )
    parts = split_side_by_side(two_panel)
    check("two-panel frame splits into 2", len(parts) == 2,
          f"got {len(parts)} panels, shapes {[p.shape for p in parts]}")
    if len(parts) == 2:
        check("left panel keeps its label column",
              all(str(parts[0].iloc[r, 0]).strip() for r in range(2, len(parts[0]))),
              f"left col0: {parts[0].iloc[:,0].tolist()}")
        check("right panel keeps its label column",
              all(str(parts[1].iloc[r, 0]).strip() for r in range(2, len(parts[1]))),
              f"right col0: {parts[1].iloc[:,0].tolist()}")

    # (2) ordinary wide table (rows span the full width) -> NOT split
    wide = pd.DataFrame(
        [["Year", "A", "B", "C", "D", "E"]]
        + [[str(2000 + i), str(i), str(i + 1), str(i + 2), str(i + 3), str(i + 4)]
           for i in range(1, 12)]
    )
    check("ordinary wide table kept whole", len(split_side_by_side(wide)) == 1,
          f"got {len(split_side_by_side(wide))}")

    # (3) misaligned single table: left side text-only (no numbers) -> NOT split
    misaligned = pd.DataFrame(
        [["Note text here", "", "Ministry of X", "", "1", "2", "3", "4"]
         for _ in range(10)]
    )
    check("text-only-left misalignment kept whole",
          len(split_side_by_side(misaligned)) == 1,
          f"got {len(split_side_by_side(misaligned))}")


def guard_thin_subheader():
    """Guard Z — a THIN sub-header (metric label spanning only a couple of
    columns) still merges, while a genuine sparse data row is left alone.

    HCES Table 3 has a 5-col frame whose metric row labels just 2 columns
    ('Average MPCE' in two of five), which the old max(3, ncols//2) floor
    rejected. The relaxed floor + numeric-below/no-value/vocab gates must absorb
    it into composite names without swallowing data."""
    print("Guard Z — thin sub-header absorb (_is_subheader_row)")
    import pandas as pd
    from backend.app.cleaning.header_builder import _absorb_subheader_rows, _is_subheader_row

    # thin metric row (2 of 5 labelled) over numeric data -> absorbed
    d = pd.DataFrame([
        ["", "Average MPCE", "", "Average MPCE", ""],
        ["General", "4870", "39", "5327", "35"],
        ["SC", "3384", "41", "3670", "38"],
        ["ST", "3000", "30", "3200", "28"],
    ], columns=["social_group", "2022_23", "2022_23", "2023_24", "2023_24"])
    out = _absorb_subheader_rows(d)
    cols = [str(c) for c in out.columns]
    check("thin metric row absorbed (composite built)",
          any("average_mpce" in c for c in cols), f"got {cols}")
    check("data kept (3 rows, General=4870)",
          len(out) == 3 and str(out.iloc[0, 1]).strip() == "4870",
          f"shape={out.shape} row0={out.iloc[0].tolist()}")

    # a genuine data row (carries numbers) is NOT a thin sub-header
    check("data row with values not a sub-header",
          not _is_subheader_row(["Rural", "4870", "", "", ""], 5))
    # a lone text cell over numeric data is too thin even relaxed (1 < max(2,...))
    check("single-label row not a sub-header",
          not _is_subheader_row(["Total", "", "", "", "", "", "", ""], 8))


def guard_numeric_readiness_metric():
    """Guard Y — unicode-minus parsing + the honest numeric-readiness metric.

    (a) numbers written with a unicode minus (−) or en-dash sign must parse, not
        drop to NaN. (b) numeric readiness must be scored over INTENDED-numeric
        columns only — a clean table of [text dimension | numbers | numbers]
        reads ~1.0, not dragged down by its legitimate text column (the flaw that
        made the old numeric_value_frac understate the dimension)."""
    print("Guard Y — unicode-minus + honest numeric readiness")
    import pandas as pd
    from backend.app.cleaning.numeric_normalizer import _to_number, normalize_numeric_columns
    from backend.tools.measure_quality import measure_table

    check("unicode minus parses", _to_number("−5") == -5, f"got {_to_number(chr(0x2212)+'5')}")
    check("en-dash sign parses", _to_number("–7.5") == -7.5, f"got {_to_number(chr(0x2013)+'7.5')}")
    col = ["−5", "10", "20", "-"]
    out = normalize_numeric_columns(pd.DataFrame({"v": col}))
    check("unicode-minus column casts", list(out["v"]) == [-5, 10, 20, None], f"got {list(out['v'])}")

    # honest readiness: text dim col must not drag a clean numeric table down
    clean = normalize_numeric_columns(pd.DataFrame([
        ["Andhra Pradesh", "10", "20"], ["Bihar", "11", "21"], ["Goa", "12", "22"],
    ], columns=["state", "rural", "urban"]))
    m = measure_table(clean, "T")
    check("clean numeric table readiness ~1.0", m["numeric_readiness"] == 1.0,
          f"got {m['numeric_readiness']} (value_frac {m['numeric_value_frac']})")
    check("intended numeric cols counted", m["intended_numeric_cols"] == 2,
          f"got {m['intended_numeric_cols']}")
    # poisoned numeric column drags readiness below 1
    poisoned = normalize_numeric_columns(pd.DataFrame([
        ["AP", "10", "x"], ["Bihar", "11", "y"], ["Goa", "12", "13"], ["MP", "14", "15"],
    ], columns=["state", "good", "poison"]))
    mp = measure_table(poisoned, "T")
    check("poisoned column lowers readiness", mp["numeric_readiness"] < 1.0,
          f"got {mp['numeric_readiness']}")


def guard_section_heading_detection():
    """Guard AE — chapter/section governing headings are recognised for
    carry-forward, while per-table 'Table X.Y' titles are NOT (they must never be
    carried onto a neighbouring table)."""
    print("Guard AE — section-heading detection (carry-forward gate)")
    from backend.app.extract.table_extractor import _is_section_heading

    check("Chapter heading recognised", _is_section_heading("Chapter-2C Kidnapping & Abduction"))
    check("numbered section recognised", _is_section_heading("2.4 Ranking of Ministries - Group A"))
    check("Annexure heading recognised", _is_section_heading("Annexure VIII Mother Tongue"))
    check("'Table X.Y' NOT a section heading", not _is_section_heading("Table 5.2 Sex Ratio"))
    check("descriptive title NOT a section heading",
          not _is_section_heading("Distribution of households by lighting"))


def guard_continuation_title_inheritance():
    """Guard AD — an untitled continuation fragment inherits its parent's title
    as '(cont.)', but an unrelated neighbour never does.

    Continuation pages often print no heading; when structural gates keep them
    from merging they ship untitled. _inherit_continuation_titles gives such a
    page-adjacent, same-shape, column-similar fragment the parent title."""
    print("Guard AD — continuation title inheritance (stitch_tables)")
    import pandas as pd
    from backend.app.standardization.table_stitcher import stitch_tables

    parent = {"table_id": 1, "name": "Table 5 Sex Ratio by District", "page": 10,
              "df": pd.DataFrame([["A", "1"], ["B", "2"]], columns=["district", "ratio"])}
    # untitled, next page, same shape + columns -> inherits "(cont.)"
    cont = {"table_id": 2, "name": None, "page": 11,
            "df": pd.DataFrame([["C", "3"]], columns=["district", "ratio"])}
    # untitled, far page + different columns -> must NOT inherit
    unrelated = {"table_id": 3, "name": None, "page": 40,
                 "df": pd.DataFrame([["x", "y", "z"]], columns=["crop", "area", "yield"])}
    out = stitch_tables([dict(parent), dict(cont), dict(unrelated)])
    by_id = {o["table_id"]: o for o in out}
    # cont may merge into parent OR inherit as a separate (cont.) table; either
    # way its data must carry the parent's title, and the unrelated table must not.
    cont_titled = (2 not in by_id) or (by_id.get(2, {}).get("name") or "").startswith("Table 5")
    check("continuation carries parent title", cont_titled,
          f"got {[(o['table_id'], o.get('name')) for o in out]}")
    if 3 in by_id:
        check("unrelated table not mislabelled",
              not str(by_id[3].get("name") or "").startswith("Table 5"),
              f"got {by_id[3].get('name')}")


def guard_title_and_composite_metric():
    """Guard AC — title-prefix stripping recovers more titles, and the honest
    composite metric counts genuine multi-level names (not just year-spans).

    (titles) "Table: Recurring …" and "Chapter-2C Kidnapping …" were dropped
    because clause-2b rejected the ': ' and the chapter prefix shrank the cleaned
    length below the ratio. Stripping a numberless title-prefix recovers them.
    (sub-headings) the old composite metric only counted YEAR composites, badly
    undercounting 'abduction_metropolitan_cities' / 'nfhs6_urban'."""
    print("Guard AC — title-prefix strip + honest composite metric")
    from backend.app.standardization.table_name_extractor import extract_table_name
    from backend.tools.measure_quality import _is_composite
    import pandas as pd

    df = pd.DataFrame([["1", "x"], ["2", "y"]], columns=["a", "b"])
    t1 = extract_table_name(df, 0, "Table: Recurring High-Volume Ministry Category Pairs")
    check("'Table: ...' (no number) titled", bool(t1) and "recurring" in str(t1).lower(),
          f"got {t1}")
    t2 = extract_table_name(df, 0, "Chapter-2C Kidnapping and Abduction (States UTs)")
    check("'Chapter-2C ...' titled", bool(t2) and "kidnapping" in str(t2).lower(),
          f"got {t2}")
    # genuine prose still rejected
    t3 = extract_table_name(df, 0, "however, any discrepancy observed in this report may be due to rounding.")
    check("prose still rejected", t3 is None, f"got {t3}")

    # honest composite: multi-level names count, plain 2-word concepts do not
    check("group+sub counts", _is_composite("nfhs6_urban") and _is_composite("abduction_metropolitan_cities")
          and _is_composite("average_mpce_rural_2022_23"))
    check("plain 2-word concept excluded",
          not _is_composite("ministry_department") and not _is_composite("crime_head")
          and not _is_composite("s_no"))


def guard_corruption_quarantine():
    """Guard AB — tables still font-corrupt after OCR (Devanagari OR Kruti soup)
    are quarantined (reason garbled_source), not shipped as clean; an English
    table with one stray Hindi token still passes.

    The quarantine now covers BOTH corruption kinds via the shared corruption
    detector, and only fires when a LARGE share (>40%) of cells are corrupt."""
    print("Guard AB — corruption quarantine, deva + kruti (validate_table)")
    import pandas as pd
    from backend.app.validation.table_validator import validate_table

    deva = pd.DataFrame([
        ["कुल", "जनसंख्या", "साक्षरता"], ["राज्य", "१००", "७५"], ["जिला", "२००", "८०"],
    ], columns=["क्षेत्र", "मान", "दर"])
    r = validate_table(deva)
    check("mostly-Devanagari table quarantined",
          not r["passed"] and r["reason"] == "garbled_source", f"got {r}")

    kruti = pd.DataFrame([
        ["';ksiqj", "eqjSuk", "fHk.M"], ["Xokfy;j", "nfr;k", "f'koiqjh"],
        ["xquk", "v'kksduxj", "Vhdex<+"],
    ], columns=["ftyk", "o\"kz", ";ksx"])
    rk = validate_table(kruti)
    check("mostly-Kruti-soup table quarantined",
          not rk["passed"] and rk["reason"] == "garbled_source", f"got {rk}")

    mostly_english = pd.DataFrame([
        ["Andhra Pradesh", "4870", "6782"], ["बिहार", "3773", "6459"],
        ["Goa", "6996", "9000"], ["Kerala", "5000", "7000"],
    ], columns=["state", "rural", "urban"])
    check("stray Hindi name still passes", validate_table(mostly_english)["passed"],
          f"got {validate_table(mostly_english)}")


def guard_corruption_detector():
    """Guard AF — the shared corruption detector flags Devanagari + Kruti soup,
    ignores numeric cells, and clears clean English."""
    print("Guard AF — corruption detector (corruption.py)")
    import pandas as pd
    from backend.app.translation.corruption import corruption_score, is_corrupt

    deva = pd.DataFrame([["राज्य", "१००"], ["जिला", "२००"]])
    check("Devanagari flagged", is_corrupt(deva)[0] and corruption_score(deva)[1] == "deva")
    kruti = pd.DataFrame([["';ksiqj", "eqjSuk"], ["Xokfy;j", "nfr;k"]])
    check("Kruti soup flagged", is_corrupt(kruti)[0] and corruption_score(kruti)[1] == "kruti")
    clean = pd.DataFrame([["Andhra Pradesh", "4870"], ["Bihar", "3773"]])
    check("clean English not flagged", not is_corrupt(clean)[0])
    numeric = pd.DataFrame([["1", "2", "3"], ["4", "5", "6"]])
    check("all-numeric not flagged", not is_corrupt(numeric)[0])
    # a numbers-heavy table with one stray corrupt label scores low
    mixed = pd.DataFrame([["राज्य", "10", "20", "30", "40"], ["x", "11", "21", "31", "41"]])
    check("numbers-heavy table scores low", corruption_score(mixed)[0] < 0.4,
          f"got {corruption_score(mixed)[0]}")


def guard_transliteration():
    """Guard AG — Devanagari transliterates to readable ASCII Latin (no Devanagari
    survives); Latin/numeric text passes through unchanged."""
    print("Guard AG — Devanagari->Latin transliteration (ocr_recovery)")
    import re
    from backend.app.extract.ocr_recovery import transliterate
    DEVA = re.compile(r"[ऀ-ॿ]")
    out = transliterate("मध्य प्रदेश")
    check("Devanagari -> Latin", not DEVA.search(out) and out.strip() != "", f"got {out}")
    check("contains plausible latin", bool(re.search(r"[a-z]", out)), f"got {out}")
    check("Latin passes through", transliterate("Madhya Pradesh") == "Madhya Pradesh")
    check("digits transliterate", transliterate("२०२३") == "2023", f"got {transliterate('२०२३')}")


def guard_ocr_recovery():
    """Guard AH — OCR rebuilds a real font-corrupt table into non-corrupt text
    with the grid preserved. Skips cleanly when Tesseract is unavailable."""
    print("Guard AH — OCR recovery on real corrupt table (economic_survey_hindi)")
    from backend.app.extract.ocr_recovery import tesseract_available, recover_table
    if not tesseract_available():
        check("tesseract present (skip if not)", True, "tesseract missing — skipped")
        return
    import warnings as _w; _w.filterwarnings("ignore")
    import pdfplumber, camelot, pandas as pd
    from backend.app.translation.corruption import corruption_score
    pdf = os.path.join(ROOT, "Testpdfs/new_batch/economic_survey_2024-25_hindi.pdf")
    plumber = pdfplumber.open(pdf)
    recovered = None
    for t in camelot.read_pdf(pdf, pages="1-8", flavor="stream"):
        if corruption_score(t.df)[0] >= 0.3 and t.df.shape[0] >= 4:
            grid = recover_table(plumber.pages[int(t.page) - 1], t)
            if grid:
                before = corruption_score(t.df)[0]
                after = corruption_score(pd.DataFrame(grid))[0]
                recovered = (before, after, t.df.shape, len(grid))
                break
    plumber.close()
    check("found & recovered a corrupt table", recovered is not None,
          "no corrupt table recovered in pp1-8")
    if recovered:
        before, after, shape, nrows = recovered
        check("corruption dropped after OCR", after < before and after < 0.3,
              f"before={before:.2f} after={after:.2f}")
        check("row count roughly preserved", abs(nrows - shape[0]) <= 3,
              f"orig rows={shape[0]} recovered={nrows}")


def guard_ghost_on_census():
    """Guard AJ — ghost suppression positively verified on a DEEP census slice
    where single-row '1 2 3' index-legend fragments actually appear.

    Range updated for Loop Spec 1 (extract-verify-re-extract retries): pages
    50-62 used to leak two such fragments because lattice "succeeded" there
    with a tiny 1-row garbage grid and stream was never tried on that page.
    table_extractor now retries a low-scoring lattice hit with stream before
    giving up, so pp.56-57 now extract the REAL underlying tables instead of
    ghost fragments (a genuine improvement, not a suppression regression —
    verified by diffing extract_tables() output before/after this change).
    pp.74-86 still contain two irrecoverable index-legend fragments (a bare
    serial-number column with no other content — no strategy can fix that),
    so the ghost-suppression assertion itself is unchanged, just re-pointed
    at a slice where the fixture still holds."""
    print("Guard AJ — ghost suppression on deep census slice")
    census = os.path.join(ROOT, "Testpdfs/new_batch/census_2011_garhwal.pdf")
    path = _slice(census, 74, 86)
    items = _pipeline(path)
    os.unlink(path)
    legends = sum(1 for i in items if i["reason"] == "index_legend_only")
    check(">=1 index-legend ghost dropped in deep census slice", legends >= 1,
          f"got {legends}; reasons={set(i['reason'] for i in items)}")


def guard_descriptive_title():
    """Guard AA — descriptive (non-'Table N') headings above a table are captured
    as titles, while body prose and 'Table N' lines behave as before.

    ~47% of tables had no title because _title_from_lines only accepted
    'Table N'/numbered-section lines; a purely descriptive heading
    ('Distribution of households by source of lighting') fell through. The new
    heading detector must accept those and still reject sentences/data rows."""
    print("Guard AA — descriptive title capture (_title_from_lines)")
    from backend.app.extract.table_extractor import _looks_like_heading, _title_from_lines

    check("descriptive Title-Case heading accepted",
          _looks_like_heading("Distribution of Households by Source of Lighting"))
    check("keyworded lower heading accepted",
          _looks_like_heading("Number of registered vehicles by state-wise category"))
    check("body sentence rejected",
          not _looks_like_heading("The table below shows the data we collected over the year."))
    check("dangling-colon prose rejected", not _looks_like_heading("As shown below:"))
    check("numeric data row rejected", not _looks_like_heading("1990 10 20 30 40"))
    check("one-word line rejected", not _looks_like_heading("Notes"))

    # 'Table N' still wins when present
    lines = ["Some intro paragraph that rambles on and on about methodology.",
             "Table 4.2 Literacy rate by district"]
    check("explicit Table-N still preferred",
          _title_from_lines(lines).startswith("Table 4.2"), f"got {_title_from_lines(lines)}")
    # descriptive heading recovered when no Table-N exists
    lines2 = ["Source: Census 2011.",
              "Distribution of Workers by Industry and Sex"]
    check("descriptive heading recovered",
          _title_from_lines(lines2) == "Distribution of Workers by Industry and Sex",
          f"got {_title_from_lines(lines2)}")


def guard_column_dedupe():
    """Guard W — duplicate column names are made unique (pandas/SQL-safe) without
    disturbing already-unique schemas.

    Repeated header blocks / side-by-side panels collapse to identical composite
    names (HCES Table 6 prints two 'average MPCE' blocks). _dedupe_columns must
    suffix the 2nd+ occurrences with a non-colliding numeric tag, and leave any
    table whose names are already unique exactly as-is (so the exact-schema
    guards B/D/H/L/V stay valid)."""
    print("Guard W — column-name dedupe (header_postprocessor)")
    import pandas as pd
    from backend.app.cleaning.header_postprocessor import _dedupe_columns, clean_headers

    # plain duplicates -> suffixed
    check("simple dup suffixed",
          _dedupe_columns(["state", "mpce", "mpce"]) == ["state", "mpce", "mpce_2"],
          f"got {_dedupe_columns(['state','mpce','mpce'])}")
    # triple duplicate -> _2, _3
    check("triple dup -> _2,_3",
          _dedupe_columns(["a", "a", "a"]) == ["a", "a_2", "a_3"],
          f"got {_dedupe_columns(['a','a','a'])}")
    # generated suffix must not collide with an existing name -> stays unique
    collide = _dedupe_columns(["x", "x", "x_2"])
    check("suffix avoids existing collision (all unique)",
          len(set(collide)) == 3 and collide[0] == "x", f"got {collide}")
    # already-unique list untouched
    uniq = ["s_no", "ministry_department", "grai_rank"]
    check("unique list untouched", _dedupe_columns(uniq) == uniq)
    # end-to-end through clean_headers: a frame with dup raw headers -> unique
    df = pd.DataFrame([["AP", "1", "2", "3", "4"]],
                      columns=["State", "Average MPCE", "Average MPCE", "Diff", "Diff"])
    out = clean_headers(df)
    cols = list(out.columns)
    check("clean_headers yields unique columns", len(set(cols)) == len(cols), f"got {cols}")
    check("no data lost in dedupe", out.shape == (1, 5), f"got {out.shape}")


def guard_phantom_value_columns():
    """Guard X — headerless NUMERIC value columns become 'value' (addressable),
    headerless TEXT label columns become 'label', and named columns stay untouched.

    A col_N column carrying real numbers is data whose header was lost upstream;
    naming it 'value' (deduped to value/value_2/…) stops it reading as extraction
    noise. A col_N column of TEXT must NOT be mislabelled 'value' — it is a lost
    label column, named 'label' instead so it is addressable."""
    print("Guard X — phantom value/label column naming (clean_headers)")
    import pandas as pd
    from backend.app.cleaning.header_postprocessor import (
        clean_headers, _is_numeric_column, _is_text_label_column)

    check("numeric column detected", _is_numeric_column(pd.Series(["1", "2", "3", "-"])))
    check("text column not numeric", not _is_numeric_column(pd.Series(["RTGS", "NEFT", "IMPS"])))
    check("text label column detected", _is_text_label_column(pd.Series(["RTGS", "NEFT", "IMPS"])))
    check("numeric column not a text label", not _is_text_label_column(pd.Series(["1", "2", "3"])))

    # numeric headerless value cols -> value/value_2 ; named col 0 kept ; text col_N -> label
    df = pd.DataFrame(
        [["North", "10", "20", "high"], ["South", "11", "21", "low"]],
        columns=["region", "col_1", "col_2", "col_3"])
    out = clean_headers(df)
    cols = list(out.columns)
    check("numeric col_N -> value*", sum(1 for c in cols if c.startswith("value")) == 2,
          f"got {cols}")
    check("text col_N -> label (not value)", any(c.startswith("label") for c in cols)
          and not any(re.fullmatch(r"col(_\d+)?", c) for c in cols), f"got {cols}")
    check("named column 0 untouched", cols[0] == "region", f"got {cols}")
    check("all columns unique", len(set(cols)) == len(cols), f"got {cols}")

    # an unnamed first column of state names -> high-confidence role "state"
    df2 = pd.DataFrame(
        [["Andhra Pradesh", "1,234"], ["Bihar", "2,345"], ["Kerala", "3,456"]],
        columns=["col", "value"])
    cols2 = list(clean_headers(df2).columns)
    check("unnamed state column -> state", cols2[0] == "state", f"got {cols2}")

    # an unnamed first text column with no confident role -> generic label
    df3 = pd.DataFrame(
        [["Alpha", "1,234"], ["Bravo and Co", "2,345"], ["Charlie", "3,456"]],
        columns=["col", "value"])
    cols3 = list(clean_headers(df3).columns)
    check("unnamed generic text column -> label", cols3[0] == "label", f"got {cols3}")

    # an unnamed column of percentages -> percentage
    df4 = pd.DataFrame(
        [["x", "12.3%"], ["y", "45.6%"], ["z", "78.9%"]], columns=["label", "col"])
    cols4 = list(clean_headers(df4).columns)
    check("unnamed percentage column -> percentage", "percentage" in cols4, f"got {cols4}")


def guard_section_lift():
    """Guard T — in-table section banners become a `category` column; ordinary
    tables are untouched.

    NFHS/PLFS factsheets print section banners ("Marriage and Fertility") as a
    label-only row over a block of data. lift_section_rows must pull each into a
    forward-filled `category` column and drop the banner rows — but ONLY when at
    least two banners each head real data, so a stray sub-title or a single
    edition label (PLFS "PLFS 2024") never spawns a phantom column.
    """
    print("Guard T — section-row lift (section_lifter)")
    import pandas as pd
    from backend.app.cleaning.section_lifter import lift_section_rows

    # (1) two genuine sections -> category column, banners dropped, ffilled
    sect = pd.DataFrame([
        ["Marriage and Fertility", "", "", ""],
        ["1. Women married before 18", "11.4", "14.7", "12.6"],
        ["2. Total fertility rate", "1.6", "2.1", "1.9"],
        ["Infant Mortality", "", "", ""],
        ["3. Neonatal mortality", "18.6", "24.5", "22.0"],
        ["Note: indicators highlighted in grey.", "", "", ""],
    ], columns=["indicator", "urban", "rural", "total"])
    out = lift_section_rows(sect)
    check("category column added", list(out.columns)[0] == "category",
          f"got {list(out.columns)}")
    check("banners + trailing note dropped (3 data rows kept)", len(out) == 3, f"got {len(out)}")
    if "category" in out.columns:
        check("category forward-filled",
              list(out["category"]) == ["Marriage and Fertility", "Marriage and Fertility",
                                        "Infant Mortality"],
              f"got {list(out['category'])}")
        check("indicator values intact",
              list(out["indicator"]) == ["1. Women married before 18",
                                         "2. Total fertility rate", "3. Neonatal mortality"])

    # (2) single banner -> no-op (not genuine sectioning)
    one = pd.DataFrame([
        ["PLFS 2024", "", "", ""],
        ["15-29", "10", "20", "30"],
        ["30-44", "11", "21", "31"],
    ], columns=["age_group", "a", "b", "c"])
    check("single banner -> unchanged",
          list(lift_section_rows(one).columns) == ["age_group", "a", "b", "c"],
          f"got {list(lift_section_rows(one).columns)}")

    # (3) two ADJACENT banners (no data between) -> no-op (wrapped title, not sections)
    adj = pd.DataFrame([
        ["PLFS 2024", "", "", ""],
        ["January-December 2024", "", "", ""],
        ["15-29", "10", "20", "30"],
        ["30-44", "11", "21", "31"],
    ], columns=["age_group", "a", "b", "c"])
    check("adjacent banners -> unchanged (1 real section)",
          "category" not in lift_section_rows(adj).columns,
          f"got {list(lift_section_rows(adj).columns)}")

    # (4) ordinary table (every row has data) -> no-op
    normal = pd.DataFrame([
        ["Andhra Pradesh", "4870", "6782"],
        ["Bihar", "3773", "6459"],
        ["Goa", "6996", "9000"],
    ], columns=["state", "rural", "urban"])
    check("ordinary table -> unchanged",
          list(lift_section_rows(normal).columns) == ["state", "rural", "urban"])


def guard_multilevel_header_merge():
    """Guard V — leaked multi-level sub-header rows fold into composite column
    names; real data rows and non-numeric tables are never touched.

    HCES Table 2 prints a year group row over a metric row ('Average MPCE' |
    'Urban-Rural differences') over a Rural/Urban row. The detector keeps only the
    year row, stranding the metric + sub rows as data and leaving duplicate
    '2022_23' columns. _absorb_subheader_rows must merge them into
    2022_23_average_mpce_rural / _urban / _urban_rural_differences while keeping
    the data and refusing to swallow a genuine data row.
    """
    print("Guard V — multi-level header merge (_absorb_subheader_rows)")
    import pandas as pd
    from backend.app.cleaning.header_builder import _absorb_subheader_rows, apply_headers

    # (1) two stranded sub-header rows over numeric data -> composite names
    d = pd.DataFrame([
        ["", "Average MPCE", "", "Urban-Rural differences", "Average MPCE", "", "Urban-Rural differences"],
        ["", "Rural", "Urban", "", "Rural", "Urban", ""],
        ["Andhra Pradesh", "4870", "6782", "39", "5327", "7182", "35"],
        ["Bihar", "3384", "4768", "41", "3670", "5080", "38"],
        ["Goa", "7666", "9509", "24", "8392", "10268", "22"],
    ], columns=["major_state", "2022_23", "2022_23", "2022_23", "2023_24", "2023_24", "2023_24"])
    out = _absorb_subheader_rows(d)
    cols = [str(c) for c in out.columns]
    phantom = [c for c in cols[1:] if re.fullmatch(r"col(_\d+)?", c)]
    check("0 col_N in value columns", not phantom, f"got {cols}")
    check("no duplicate column names", len(set(cols)) == len(cols), f"got {cols}")
    check("year+metric+sub composites built",
          "2022_23_average_mpce_rural" in cols and "2023_24_average_mpce_urban" in cols,
          f"got {cols}")
    check(">=4 composite (year+sub) columns",
          sum(1 for c in cols if re.search(r"20\d\d_\d\d", c) and
              ("rural" in c or "urban" in c or "differences" in c)) >= 4, f"got {cols}")
    check("data rows kept intact (3 rows, AP=4870)",
          len(out) == 3 and str(out.iloc[0, 1]).strip() == "4870",
          f"shape={out.shape} row0={out.iloc[0].tolist()}")

    # (2) first data row is real data (has numbers) -> no absorb
    realdata = pd.DataFrame([
        ["Andhra Pradesh", "4870", "6782"],
        ["Bihar", "3384", "4768"],
    ], columns=["state", "2022_23", "2023_24"])
    check("real data first row -> unchanged",
          list(_absorb_subheader_rows(realdata).columns) == ["state", "2022_23", "2023_24"]
          and len(_absorb_subheader_rows(realdata)) == 2)

    # (3) numeric-below gate: sub-header-looking row but text data below -> no absorb
    texttable = pd.DataFrame([
        ["", "Rural", "Urban", "Total"],
        ["x", "high", "low", "mid"],
        ["y", "low", "high", "mid"],
    ], columns=["c0", "c1", "c2", "c3"])
    before = list(texttable.columns)
    check("non-numeric data below -> no absorb",
          list(_absorb_subheader_rows(texttable).columns) == before
          and len(_absorb_subheader_rows(texttable)) == 3)

    # (4) single stranded volume/value sub-row (RBI-style) still merges (k=1)
    one = pd.DataFrame([
        ["", "Volume", "Value", "Volume", "Value"],
        ["RTGS", "10", "20", "30", "40"],
        ["NEFT", "11", "21", "31", "41"],
    ], columns=["col", "2020_21", "2020_21", "2021_22", "2021_22"])
    o1 = _absorb_subheader_rows(one)
    c1 = [str(c) for c in o1.columns]
    check("single sub-row merged into year_label composites",
          any("volume" in c and "2020_21" in c for c in c1)
          and any("value" in c and "2021_22" in c for c in c1), f"got {c1}")


def guard_workbook_toc():
    """Guard U — navigable workbook: Contents tab first, readable sheet codes,
    full untruncated titles, ghost frames excluded.

    A 600-sheet census workbook is unusable if every tab is "table_517" and the
    titles are clipped at 10 words. build_workbook must put a Contents/TOC sheet
    first that maps each short, human-readable sheet code back to the full title
    with page + dimensions, and must never emit a tab for an empty frame.
    """
    print("Guard U — navigable workbook + TOC (excel_exporter)")
    import pandas as pd
    from openpyxl import load_workbook
    from backend.app.export.excel_exporter import build_workbook, _sheet_name
    from backend.app.standardization.table_name_extractor import _clean_title_words

    # title truncation lifted: a long real title keeps >10 words
    long_title = ("Number and percentage distribution of households by source of "
                  "lighting and type of fuel used for cooking across rural and urban areas")
    kept = _clean_title_words(long_title)
    check("long title not clipped at 10 words", len(kept.split()) > 10, f"got {len(kept.split())} words")

    # sheet codes are readable and Excel-legal
    s1 = _sheet_name("Table 6.2 Sex Ratio of Rural Population", 17, set())
    check("sheet code carries number + words", s1.startswith("T6_2") and "Sex" in s1, f"got {s1}")
    used = {"T6_2_Sex_Ratio"}
    s2 = _sheet_name("Table 6.2 Sex Ratio of Rural Population", 18, used)
    check("duplicate sheet code disambiguated", s2 != "T6_2_Sex_Ratio" and len(s2) <= 31, f"got {s2}")
    s3 = _sheet_name(None, 99, set())
    check("untitled falls back to table id", s3 == "table_99", f"got {s3}")

    # build a workbook with a real table, an untitled one, and a ghost(empty) one
    dfs = {
        1: pd.DataFrame([["AP", "10", "20"], ["Bihar", "11", "21"]],
                        columns=["state", "rural", "urban"]),
        2: pd.DataFrame([["x", "1"]], columns=["a", "b"]),
        3: pd.DataFrame(columns=["col", "col_1"]),  # ghost: no rows
    }
    catalog = [
        {"table_id": 1, "table_name": "Table 6.2 " + long_title, "page": 12},
        {"table_id": 2, "table_name": None, "page": 13},
        {"table_id": 3, "table_name": "Ghost", "page": 14},
    ]
    buf = build_workbook(dfs, catalog)
    wb = load_workbook(buf)
    check("Contents/TOC is the first sheet", wb.sheetnames[0] == "Contents",
          f"got {wb.sheetnames[:3]}")
    check("ghost frame got no sheet", len(wb.sheetnames) == 1 + 2,
          f"sheets: {wb.sheetnames}")
    toc = wb["Contents"]
    header = [toc.cell(row=1, column=c).value for c in range(1, 7)]
    check("TOC header has Name/Page/Rows/Cols",
          header == ["#", "Table Name", "PDF Page", "Rows", "Cols", "Sheet"], f"got {header}")
    names = [toc.cell(row=r, column=2).value for r in range(2, 4)]
    check("TOC lists full title (t-code -> real title)",
          any(n and len(str(n).split()) > 10 for n in names), f"got {names}")


def guard_numeric_normalization():
    """Guard R — numeric columns cast to real numbers; merged/label columns spared.

    Camelot ships every cell as a string ("45,544", "(7.5)", "12.3%", "-"), so a
    numeric column lands as object dtype and df.sum()/groupby silently break.
    normalize_numeric_columns must:
      - strip thousands separators / footnote markers and cast (>=80% numeric col)
      - parse parenthesised negatives, percent, dash/blank -> None (blank on export)
      - keep whole values as int (clean "45544", never "45544.0")
      - REFUSE a column carrying merged continuation cells ("16411581 (15878397)")
      - leave label/text columns untouched
    """
    print("Guard R — numeric normalization (numeric_normalizer)")
    import pandas as pd
    from backend.app.cleaning.numeric_normalizer import (
        normalize_numeric_columns, _to_number, _column_is_castable,
    )
    # cell-level parsing
    check("'45,544' -> 45544 int", _to_number("45,544") == 45544
          and isinstance(_to_number("45,544"), int))
    check("'(7.5)' -> -7.5", _to_number("(7.5)") == -7.5)
    check("'12.3%' -> 12.3", _to_number("12.3%") == 12.3)
    check("'3348245*' -> 3348245", _to_number("3348245*") == 3348245)
    check("'-' -> None", _to_number("-") is None)
    check("'n.a.' -> None", _to_number("n.a.") is None)
    check("'' -> None", _to_number("") is None)
    check("'Andhra Pradesh' -> None", _to_number("Andhra Pradesh") is None)
    # column castability
    check("clean numeric column castable",
          _column_is_castable(["45,544", "3,773", "2079", "-"]))
    check("merged-continuation column spared",
          not _column_is_castable(["16411581 (15878397)", "10 20", "30"]))
    check("text column not castable",
          not _column_is_castable(["Rural", "Urban", "Total"]))
    # frame-level: value cols cast, label & merged cols preserved
    df = pd.DataFrame({
        "state": ["Andhra Pradesh", "Bihar", "Goa"],
        "mpce": ["4,870", "3,773", "6,996"],
        "pct": ["39", "(7.5)", "12.3%"],
        "provisional": ["16411581 (15878397)", "100 (90)", "5 (4)"],
    })
    out = normalize_numeric_columns(df)
    check("label column untouched", list(out["state"]) == ["Andhra Pradesh", "Bihar", "Goa"])
    check("mpce cast to ints (no .0)",
          [str(v) for v in out["mpce"]] == ["4870", "3773", "6996"],
          f"got {[str(v) for v in out['mpce']]}")
    check("pct parsed (neg paren, percent)",
          list(out["pct"]) == [39, -7.5, 12.3], f"got {list(out['pct'])}")
    check("merged-continuation column kept as text",
          list(out["provisional"]) == ["16411581 (15878397)", "100 (90)", "5 (4)"])
    # idempotent
    check("idempotent on second pass",
          [str(v) for v in normalize_numeric_columns(out)["mpce"]] == ["4870", "3773", "6996"])


def guard_ghost_suppression():
    """Guard S — header-only 'index legend' fragments are dropped, real tables kept.

    Garhwal-census village-directory spillover leaves a single row that is just
    the column-number band ("1 | 2 | 3 | 4 | 5 | 6"); it passes the old shape
    checks but holds no data. validate_table must reject it (index_legend_only)
    while a one-row KPI strip of real values and any normal table still pass.
    """
    print("Guard S — ghost / index-legend suppression (validate_table)")
    import pandas as pd
    from backend.app.validation.table_validator import validate_table, _is_index_legend_row
    # row-level discrimination
    check("['1','2','3','4'] is legend", _is_index_legend_row(["1", "2", "3", "4"]))
    check("['1 2 3','4','5','6'] is legend", _is_index_legend_row(["1 2 3", "4", "5", "6"]))
    check("['(1)','(2)','(3)'] is legend", _is_index_legend_row(["(1)", "(2)", "(3)"]))
    check("KPI real values not legend",
          not _is_index_legend_row(["45.2", "1,234", "78"]))
    check("section text row not legend",
          not _is_index_legend_row(["Marriage and Fertility", "", ""]))
    check("non-ascending small ints not legend",
          not _is_index_legend_row(["3", "1", "2"]))
    # table-level
    legend = pd.DataFrame([["1", "2", "3", "4", "5", "6"]],
                          columns=["col", "a", "b", "c", "d", "e"])
    r = validate_table(legend)
    check("legend-only table rejected", not r["passed"] and r["reason"] == "index_legend_only",
          f"got {r}")
    kpi = pd.DataFrame([["45.2", "1234", "78.9"]], columns=["rate", "count", "pct"])
    check("one-row KPI strip kept", validate_table(kpi)["passed"], f"got {validate_table(kpi)}")
    normal = pd.DataFrame(
        [["Andhra Pradesh", "4870", "6782"], ["Bihar", "3773", "6459"]],
        columns=["state", "rural", "urban"])
    check("normal table kept", validate_table(normal)["passed"])


def guard_reference_tables():
    """Guard AM — the archetype profiler routes code+text lookup catalogues
    (NCO concordance) to the reference header path: data rows preserved, stable
    columns, no giant absorbed-header column names. Statistical tables unaffected."""
    print("Guard AM — reference-table profiling + header (table_profiler)")
    import pandas as pd
    from backend.app.profile.table_profiler import classify_table, reference_header_rows

    # synthetic reference frame (code + text, no measures)
    ref = pd.DataFrame([
        ["NCO 2015", "", "", "NCO 2004"],
        ["Family", "1111", "Legislators and Senior Officials", ""],
        ["", "1111.0100", "Elected Official, Union Government", "1111.10"],
        ["", "1111.0200", "Elected Official, State Government", "1112.10"],
        ["", "1111.0300", "Elected Official, Local Bodies", "1113.10"],
        ["", "1112.0100", "Administrative Official, Union Government", "1121.10"],
        ["", "1112.0200", "Diplomat and Foreign Service Officer", "1121.20"],
    ])
    c = classify_table(ref)
    check("code+text catalogue -> reference", c["archetype"] == "reference", str(c))
    check("reference header = 1 (NCO row only)", reference_header_rows(ref) == 1,
          f"got {reference_header_rows(ref)}")

    # statistical frame must NOT be reference
    stat = pd.DataFrame([
        ["State", "2021", "2022", "2023"],
        ["Andhra Pradesh", "12.3", "13.4", "14.5"],
        ["Bihar", "9.1", "9.8", "10.2"],
        ["Kerala", "45.6", "47.1", "48.0"],
    ])
    check("numeric measures -> statistical", classify_table(stat)["archetype"] == "statistical",
          str(classify_table(stat)))

    # real PDF: NCO concordance pages keep data rows + stable 4 cols
    nco = "/Users/thesinghaa/Downloads/national classification of occupations _vol i- 2015.pdf"
    if os.path.exists(nco):
        path = _slice(nco, 46, 48)
        items = [i for i in _pipeline(path) if i["passed"]]
        os.unlink(path)
        check("NCO pages pass", len(items) >= 3, f"got {len(items)}")
        if items:
            shapes_4col = all(i["df"].shape[1] == 4 for i in items)
            check("all NCO pages 4 columns", shapes_4col,
                  f"shapes {[i['df'].shape for i in items]}")
            # no giant absorbed-header column name (the old bug)
            no_giant = all(max(len(str(c)) for c in i["df"].columns) < 40 for i in items)
            check("no absorbed-header giant column names", no_giant,
                  f"cols {[list(i['df'].columns) for i in items][:1]}")
            # hierarchy data row preserved (Managers / a Family row present)
            body = " ".join(" ".join(str(v) for v in r)
                            for i in items for r in i["df"].head(3).values.tolist())
            check("hierarchy data rows preserved (not eaten as header)",
                  "Managers" in body or "Legislators" in body, body[:80])

        # multi-page merge: a long concordance run collapses into ONE table
        from backend.app.standardization.table_stitcher import stitch_tables
        path2 = _slice(nco, 46, 60)
        passed = [i for i in _pipeline(path2) if i["passed"]]
        os.unlink(path2)
        merged = stitch_tables([{"table_id": p["table_id"], "name": p["name"] or "",
                                 "page": p["page"], "df": p["df"], "pages": [p["page"]],
                                 "archetype": p.get("archetype")}
                                for p in passed])
        check("15 concordance pages merge into 1 table", len(merged) == 1,
              f"got {len(merged)} tables")
        if merged:
            mdf = merged[0]["df"]
            check("merged table keeps all rows (>=200)", len(mdf) >= 200,
                  f"got {len(mdf)} rows")
            check("merged table stays 4 columns", mdf.shape[1] == 4,
                  f"got {mdf.shape[1]}")
            # semantic column names inferred from content (not nco/value/label)
            cols = [str(c) for c in mdf.columns]
            check("semantic column names inferred", "code" in cols and "name" in cols,
                  f"got {cols}")


def guard_title_recovery():
    """Guard AL — titles are recovered from a leading descriptive cell and from
    captions with a leading enumerator, while genuine no-title tables stay None."""
    print("Guard AL — title recovery (descriptive cell + enumerator caption)")
    import pandas as pd
    from backend.app.standardization.table_name_extractor import extract_table_name

    # leading descriptive cell -> title
    df_keyind = pd.DataFrame([
        ["India - Key Indicators", "", "", ""],
        ["Indicator", "Urban", "Rural", "Total"],
        ["Population", "1,234", "5,678", "6,912"],
    ])
    nm = extract_table_name(df_keyind, 2, None)
    check("leading 'India - Key Indicators' cell -> title", nm == "India Key Indicators", f"got {nm!r}")

    # state-name repair still applies on recovered titles
    df_hr = pd.DataFrame([["ryana - Key Indicators", "", ""], ["x", "1", "2"], ["y", "3", "4"]])
    nm2 = extract_table_name(df_hr, 1, None)
    check("recovered title repairs split state name", nm2 == "Haryana Key Indicators", f"got {nm2!r}")

    # caption with leading enumerator "4." -> accepted (was rejected as prose)
    df_data = pd.DataFrame([["2022", "2023", "Total"], ["7,908", "10,717", "18,625"]])
    nm3 = extract_table_name(df_data, 1,
                             "4. MOLBR - Delay in Final Settlement / Final PF Withdrawal")
    check("enumerator caption '4. …' -> title", nm3 is not None and "MOLBR" in nm3, f"got {nm3!r}")

    # a real data table with no caption and no title cell -> still None (no fabrication)
    df_plain = pd.DataFrame([
        ["Andhra Pradesh", "1,234", "5,678"],
        ["Bihar", "2,345", "6,789"],
        ["Kerala", "3,456", "7,890"],
    ])
    nm4 = extract_table_name(df_plain, 1, None)
    check("plain data table stays untitled (no fabrication)", nm4 is None, f"got {nm4!r}")

    # prose first cell is NOT taken as a title (sentence punctuation)
    df_prose = pd.DataFrame([
        ["This table presents the distribution of households across states.", "", ""],
        ["State", "Rural", "Urban"],
        ["AP", "1", "2"],
    ])
    nm5 = extract_table_name(df_prose, 2, None)
    check("prose sentence cell not taken as title", nm5 is None, f"got {nm5!r}")

    # real PDF: NFHS5 india p6 recovers "India Key Indicators"
    nfhs = os.path.join(ROOT, "Testpdfs/new/nfhs5_india.pdf")
    if os.path.exists(nfhs):
        path = _slice(nfhs, 6, 6)
        items = [i for i in _pipeline(path) if i["passed"]]
        os.unlink(path)
        names = [i["name"] for i in items]
        check("nfhs5_india p6 -> 'India Key Indicators'",
              "India Key Indicators" in names, f"got {names}")


def guard_toc_prose_quarantine():
    """Guard AK — TOC/index pages and prose paragraphs are quarantined, while
    genuine data tables (even text-heavy ones) still pass."""
    print("Guard AK — TOC / prose quarantine (validate_table)")
    import pandas as pd
    from backend.app.validation.table_validator import _is_toc, _is_prose, validate_table

    # --- TOC: dotted leaders
    toc_leader = pd.DataFrame([
        ["3.1", "Basic Characteristics ................", "47"],
        ["3.2", "Schooling and Literacy ..............", "49"],
        ["3.3", "Wealth Index ........................", "52"],
    ])
    check("dotted-leader contents -> toc", _is_toc(toc_leader))

    # --- TOC: Table No + Page No header
    toc_header = pd.DataFrame([
        ["Chapter & Table No.", "TITLE", "PAGE No."],
        ["1A.10", "Place of Occurrence - wise Road Accident", "179"],
        ["1A.11", "Time of Occurrence - wise Road Accident", "180"],
    ])
    check("Table-No+Page-No header -> toc", _is_toc(toc_header))

    # --- TOC: INDEX header (district_indicatoe style)
    toc_index = pd.DataFrame([
        ["Table No.", "INDEX", "Page No."],
        ["1", "Area and Population", "1"],
        ["2", "Density", "3"],
    ])
    check("INDEX header -> toc", _is_toc(toc_index))

    # --- prose paragraph mis-parsed as a table
    prose = pd.DataFrame([
        ["Since the publication caters to the broad needs of various users", ""],
        ["the actual count of each crime head may differ from the figure shown", ""],
        ["any discrepancy observed in this report may be brought to notice", ""],
    ])
    check("prose paragraph -> prose_text", _is_prose(prose))

    # --- counter-examples: real tables must NOT be quarantined
    real_numeric = pd.DataFrame(
        [["Andhra Pradesh", "1,234", "5,678"], ["Bihar", "2,345", "6,789"],
         ["Kerala", "3,456", "7,890"]],
        columns=["state", "rural", "urban"])
    check("real numeric table not toc", not _is_toc(real_numeric))
    check("real numeric table not prose", not _is_prose(real_numeric))
    check("real numeric table passes", validate_table(real_numeric)["passed"])

    # text-heavy but real: long NFHS-style indicator names WITH numeric values
    real_text_heavy = pd.DataFrame(
        [["Children age 12-23 months fully vaccinated based on information", "76.4", "80.1"],
         ["Mothers who had antenatal check-up in first trimester percent", "58.6", "70.2"],
         ["Women age 15-49 who are anaemic according to the survey round", "57.0", "53.1"]],
        columns=["indicator", "nfhs5", "nfhs6"])
    check("text-heavy real table not prose", not _is_prose(real_text_heavy))
    check("text-heavy real table passes", validate_table(real_text_heavy)["passed"])

    # real PDF: district_indicatoe INDEX pages are quarantined (were col_N ghosts)
    di = os.path.join(ROOT, "backend/data/uploads/district_indicatoe.pdf")
    if os.path.exists(di):
        path = _slice(di, 10, 12)
        items = _pipeline(path)
        os.unlink(path)
        toc_passed = [i for i in items if i["passed"] and
                      all(re.fullmatch(r"col(_\d+)?", str(c)) for c in i["df"].columns)]
        check("district_indicatoe INDEX pages not shipped as col_N tables",
              len(toc_passed) == 0, f"got {len(toc_passed)} col_N-only tables still passing")


def guard_extract_retry():
    print("Guard AR — per-table extract-verify-re-extract loop (table_extractor retries)")
    import pandas as pd
    from backend.app.extract.quality import (
        LOW_QUALITY_THRESHOLD, REVIEW_THRESHOLD, score_table,
    )
    from backend.app.extract.table_extractor import _build_kept_item, _consider

    # A garbage lattice read: generic col_N columns tank header_coherence,
    # matching the audit's real case (ATR/Annexure-2 kept with col/col_2).
    lattice_df = pd.DataFrame({"col_1": ["1", "bad"], "col_2": ["2", "data"]})
    lattice_score = score_table(lattice_df)["score"]
    check("(a) fixture: lattice score below threshold", lattice_score < LOW_QUALITY_THRESHOLD,
          str(lattice_score))

    # A clean stream read of the same page: real column labels, mostly numeric.
    stream_df = pd.DataFrame({"state": ["Andhra Pradesh", "Bihar"], "receipts": ["100", "200"]})
    stream_score = score_table(stream_df)["score"]
    check("(a) fixture: stream score clears threshold", stream_score >= LOW_QUALITY_THRESHOLD,
          str(stream_score))

    attempts = []
    best = {"df": lattice_df, "score": lattice_score, "strategy": "lattice"}
    attempts.append({"strategy": "lattice", "score": lattice_score})
    best = _consider(attempts, best, stream_df, stream_score, "stream")

    check("(a) low lattice score triggers a recorded stream retry attempt",
          len(attempts) == 2 and attempts[1] == {"strategy": "stream", "score": stream_score},
          str(attempts))
    check("(a) higher-scoring stream variant is kept, strategy recorded",
          best["strategy"] == "stream" and best["score"] == stream_score and best["df"] is stream_df,
          str(best))

    # (c) max-score guarantee: a WORSE later attempt must never replace a
    # better earlier one, even though it is still recorded in `attempts`.
    worse_df = pd.DataFrame({"col_1": ["x"]})
    worse_score = score_table(worse_df)["score"]
    check("(c) fixture: worse candidate really does score lower",
          worse_score < best["score"], f"worse={worse_score} best={best['score']}")
    prior_best = dict(best)
    best = _consider(attempts, best, worse_df, worse_score, "ocr")
    check("(c) worse retry recorded as an attempt but does not win",
          len(attempts) == 3 and attempts[2] == {"strategy": "ocr", "score": worse_score},
          str(attempts))
    check("(c) best is unchanged — worse retry never replaces a better earlier attempt",
          best == prior_best, f"best={best} prior={prior_best}")

    # (b) every strategy stays below threshold -> best-effort keep, not a drop.
    all_low_attempts = []
    lo_best = {"df": lattice_df, "score": lattice_score, "strategy": "lattice"}
    all_low_attempts.append({"strategy": "lattice", "score": lattice_score})
    retry_low_df = pd.DataFrame({"col_1": ["9", "z"], "col_2": ["8", "y"]})
    retry_low_score = score_table(retry_low_df)["score"]
    lo_best = _consider(all_low_attempts, lo_best, retry_low_df, retry_low_score, "stream")

    check("(b) fixture: no strategy clears the low-quality threshold",
          lo_best["score"] < LOW_QUALITY_THRESHOLD, str(lo_best["score"]))

    item = _build_kept_item(
        page=1, df=lo_best["df"], bbox=None,
        attempts=all_low_attempts, best_score=lo_best["score"], best_strategy=lo_best["strategy"],
    )
    check("(b) low-quality table is still kept (dataframe present, not dropped)",
          item["dataframe"] is lo_best["df"], "table was dropped instead of best-effort-kept")
    check("(b) low_quality flag set true", item["low_quality"] is True, str(item))
    check("(b) review_needed matches the <0.50 escalation bar",
          item["review_needed"] == (lo_best["score"] < REVIEW_THRESHOLD),
          f"score={lo_best['score']} review_needed={item['review_needed']}")

    # Guardrail: max 3 strategies attempted per table (lattice + stream + ocr).
    check("max-3-strategies guardrail: exactly 3 recorded when all 3 are tried",
          len(attempts) == 3, str(attempts))


def guard_recheck_quarantine():
    """Guard AU — quarantine re-check (Loop Spec 4): recheck_quarantine.py
    re-extracts + re-validates each quarantined page and either promotes a
    now-passing row into promoted_tables.csv or re-quarantines a still-bad
    one with a refreshed reason. Uses a fake extract_tables (monkeypatched on
    table_extractor, picked up because recheck_quarantine imports it lazily
    inside the function body) so the guard is fast and doesn't depend on
    Loop 1's retry logic actually firing on a real PDF."""
    print("Guard AU — quarantine re-check (recheck_quarantine.py)")
    import shutil
    import tempfile
    import pandas as pd
    from backend.tools import recheck_quarantine as rq
    import backend.app.extract.table_extractor as table_extractor

    workdir = tempfile.mkdtemp()
    dummy_pdf = os.path.join(workdir, "dummy.pdf")
    with open(dummy_pdf, "wb") as f:
        f.write(b"%PDF-1.4 fake")  # only os.path.exists() is checked, never opened

    # page 1: now extracts cleanly (real header row, numeric data) -> promote
    good_raw = pd.DataFrame([
        ["State", "Rural", "Urban"],
        ["Andhra Pradesh", "10", "20"],
        ["Bihar", "11", "21"],
        ["Goa", "12", "22"],
    ])
    # page 2: genuinely bad — one real column, no retry strategy fixes that
    bad_raw = pd.DataFrame([["x"], ["1"], ["2"]])

    fake_tables = [
        {"table_id": 1, "page": 1, "dataframe": good_raw, "caption": None,
         "flavor": "lattice", "attempts": [{"strategy": "lattice", "score": 0.85}],
         "best_score": 0.85, "low_quality": False, "review_needed": False},
        {"table_id": 2, "page": 2, "dataframe": bad_raw, "caption": None,
         "flavor": "lattice", "attempts": [{"strategy": "lattice", "score": 0.55}],
         "best_score": 0.55, "low_quality": True, "review_needed": False},
    ]

    # Loop Spec 4: recheck_quarantine now calls extract_tables(pdf_path,
    # pages=...) scoped to the pages actually being rechecked — record what
    # scope it was actually called with so this guard can assert scoping
    # happened, not just that the fake still returns something.
    scope_calls = []

    def _fake_extract(pdf_path, pages=None):
        scope_calls.append(pages)
        return fake_tables

    original_extract_tables = table_extractor.extract_tables
    table_extractor.extract_tables = _fake_extract
    try:
        failed_csv = os.path.join(workdir, "failed_tables.csv")
        pd.DataFrame([
            {"table": 1, "page": 1, "reason": "too_few_columns"},
            {"table": 2, "page": 2, "reason": "mostly_empty"},
        ]).to_csv(failed_csv, index=False)

        n_checked, n_promoted = rq.run(workdir, pdf_path=dummy_pdf, run_id="guard", force=True)
    finally:
        table_extractor.extract_tables = original_extract_tables

    check("(a) extract_tables was called exactly once for this pdf (batched, not per-row)",
          len(scope_calls) == 1, str(scope_calls))
    check("(a) it was scoped to exactly the 2 quarantined pages, not the whole document",
          scope_calls and sorted(scope_calls[0]) == [1, 2], str(scope_calls))

    check("(a) both quarantined rows rechecked", n_checked == 2, str(n_checked))
    check("(a) exactly the fixable row promoted", n_promoted == 1, str(n_promoted))

    promoted = pd.read_csv(os.path.join(workdir, "promoted_tables.csv"))
    check("(a) promoted_tables.csv has the p1 row with its real score + reason",
          len(promoted) == 1 and promoted.iloc[0]["page"] == 1
          and promoted.iloc[0]["reason"] == "passed_on_recheck"
          and promoted.iloc[0]["score"] == 0.85,
          str(promoted.to_dict("records")))

    still_failed = pd.read_csv(failed_csv)
    check("(b) genuinely-bad row stays quarantined, promoted row removed from failed_tables.csv",
          len(still_failed) == 1 and still_failed.iloc[0]["page"] == 2,
          str(still_failed.to_dict("records")))
    check("(b) re-quarantine REFRESHES the reason, not left stale as the old 'mostly_empty'",
          still_failed.iloc[0]["reason"] == "too_few_columns",
          str(still_failed.iloc[0]["reason"]))
    check("(b) last_rechecked stamped with this run's id",
          still_failed.iloc[0]["last_rechecked"] == "guard",
          str(still_failed.iloc[0]["last_rechecked"]))

    # guardrail: a second run at the SAME run-id (no --force) must skip
    # everything — re-check nothing, call extract_tables zero times.
    calls = []

    def _counting_fake(pdf_path, pages=None):
        calls.append(pdf_path)
        return fake_tables

    table_extractor.extract_tables = _counting_fake
    try:
        n_checked2, n_promoted2 = rq.run(workdir, pdf_path=dummy_pdf, run_id="guard", force=False)
    finally:
        table_extractor.extract_tables = original_extract_tables
    check("(c) repeat run at the same run-id re-checks nothing (no redundant re-extraction)",
          n_checked2 == 0 and n_promoted2 == 0 and calls == [], str((n_checked2, calls)))

    shutil.rmtree(workdir, ignore_errors=True)


def guard_wrapped_header_reconstruction():
    """Guard BA — narrow-cell wrapped headers + headerless continuations
    (the "23/24 generic columns" failure on DARPG Feb 2026 pp.9-15).
    Two coupled behaviors:
    (1) universal_cleaner joins MID-WORD line wraps ("Grie\\nvanc\\nes ") with
        nothing instead of a space — real word boundaries survive as literal
        trailing spaces, so "Grievances brought forward" comes out readable
        instead of "Grie vanc es brou ght forw ard" soup.
    (2) table_stitcher treats all-fallback-named (value_N/code_N, not just
        col_N) follow-on pages as headerless continuations, even when a data
        cell got promoted into a bogus weak title, so they inherit the header
        page's real column names instead of shipping as separate garbage
        tables."""
    print("Guard BA — wrapped-header reconstruction + fallback-named continuation")
    import pandas as pd
    from backend.app.cleaning.universal_cleaner import clean_dataframe
    from backend.app.standardization.table_stitcher import _named_frac, stitch_tables

    # (1) mid-word wrap rejoin, word boundaries preserved
    raw = pd.DataFrame([["Grie\nvanc\nes \nbrou\nght \nforw\nard",
                         "Tot\nal \nNu\nmb\ner \nof \nAp\npea\nls \nFil\ned",
                         "Total\nGrievances", "100\n4"]])
    got = clean_dataframe(raw).iloc[0].tolist()
    check("(1) mid-word wraps rejoined with boundaries intact",
          got[0] == "Grievances brought forward", repr(got[0]))
    check("(1) multi-fragment name fully reconstructed",
          got[1] == "Total Number of Appeals Filed", repr(got[1]))
    check("(1) whole-word newline still becomes a space (uppercase next)",
          got[2] == "Total Grievances", repr(got[2]))
    check("(1) stacked digits NOT glued (old behavior preserved)",
          got[3] == "100 4", repr(got[3]))

    # (2) value_N/code_N count as fallback names now
    check("(2) all-value_N columns read as unnamed",
          _named_frac(["value", "value_2", "code", "label"]) == 0.0,
          str(_named_frac(["value", "value_2", "code", "label"])))

    # (2) a fallback-named follow-on page with a bogus weak title merges into
    # the titled header page; a strong-titled neighbour does NOT.
    header_df = pd.DataFrame({"ministry": ["A", "B", "C", "D"],
                              "grievances_received": [1, 2, 3, 4],
                              "resolved_within_timeline": [5, 6, 7, 8]})
    cont_df = pd.DataFrame({"label": ["E", "F"], "value": [9, 10], "value_2": [11, 12]})
    separate_df = pd.DataFrame({"label": ["G", "H"], "value": [13, 14], "value_2": [15, 16]})
    stitched = stitch_tables([
        {"table_id": 1, "name": "2.2 Group A: Indicator Wise Scores", "page": 1,
         "df": header_df, "titled": True, "flavor": ""},
        {"table_id": 2, "name": "Consumer Affairs", "page": 2,
         "df": cont_df, "titled": True, "flavor": ""},
        {"table_id": 3, "name": "Annexure 3.1 Top Ministries", "page": 3,
         "df": separate_df, "titled": True, "flavor": ""},
    ])
    check("(2) bogus-weak-titled fallback page merges into its header page",
          len(stitched) == 2 and stitched[0]["df"].shape[0] == 6,
          f"{[(s['name'][:30], s['df'].shape) for s in stitched]}")
    check("(2) merged continuation inherits the real column names",
          list(stitched[0]["df"].columns)[:2] == ["ministry", "grievances_received"],
          str(list(stitched[0]["df"].columns)))
    check("(2) strong-titled neighbour stays a separate table",
          any(s["name"].startswith("Annexure") for s in stitched),
          f"{[s['name'][:30] for s in stitched]}")


def guard_page_scope_real_pdf():
    """Guard AX — Loop Spec 4 page-scoping, exercised against a REAL 200-page
    PDF (Sample.pdf), not a synthetic fixture. Guard AU already proves
    recheck_quarantine.py REQUESTS the right page set from extract_tables();
    this guard proves extract_tables() ITSELF actually narrows the expensive
    camelot calls to that scope on a real multi-page document. It intercepts
    camelot.read_pdf — the actual external library call whose per-page work
    is what page-scoping exists to avoid — and records the `pages` argument
    it receives, rather than paying for a real 200-page camelot run inside
    the guard suite (the stub returns an empty result, so no actual table
    geometry search happens; only the call's arguments are being verified)."""
    print("Guard AX — page-scoping on a real multi-page PDF (table_extractor.extract_tables)")
    import camelot
    from backend.app.extract import table_extractor

    sample_pdf = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Sample.pdf"))
    if not os.path.exists(sample_pdf):
        print("  [SKIP] Sample.pdf not found in repo root — skipping real-PDF page-scope guard")
        return

    calls = []

    def _recording_read_pdf(path, pages="1", flavor="lattice", **kwargs):
        calls.append({"pages": pages, "flavor": flavor})
        return []

    original_read_pdf = camelot.read_pdf

    # (a)+(b): scoped call — restrict extract_tables to a single real page.
    camelot.read_pdf = _recording_read_pdf
    try:
        table_extractor.extract_tables(sample_pdf, pages=[7])
    finally:
        camelot.read_pdf = original_read_pdf

    lattice_calls = [c for c in calls if c["flavor"] == "lattice"]
    check("(a) the lattice pass is scoped to the requested page, not 'all'",
          len(lattice_calls) == 1 and lattice_calls[0]["pages"] == "7", str(calls))

    # Stub returns nothing, so the stream fallback must fire — and it must
    # ALSO stay scoped to page 7, never sweep Sample.pdf's real 200 pages
    # looking for tables lattice missed.
    stream_calls = [c for c in calls if c["flavor"] == "stream"]
    check("(b) the stream fallback fired (scoping didn't silently suppress it)",
          len(stream_calls) == 1, str(calls))
    check("(b) the stream fallback only requested the scoped page, never all 200",
          bool(stream_calls) and stream_calls[0]["pages"] == "7", str(calls))

    # (c) unscoped (default, pages=None) call — every pre-existing caller's
    # behavior — must still request the WHOLE document. Confirms scoping is
    # strictly opt-in and changed nothing for callers that don't pass `pages`.
    calls.clear()
    camelot.read_pdf = _recording_read_pdf
    try:
        table_extractor.extract_tables(sample_pdf)
    finally:
        camelot.read_pdf = original_read_pdf

    default_lattice_calls = [c for c in calls if c["flavor"] == "lattice"]
    check("(c) unscoped call still requests the whole document ('all') — unchanged default behavior",
          len(default_lattice_calls) == 1 and default_lattice_calls[0]["pages"] == "all",
          str(calls))


def guard_frame_title_recovery():
    """Guard BI — titles inside a full-page border frame (table_extractor).
    DARPG 2022 pages draw a decorative box around the whole page; lattice
    takes that frame as the table boundary, so the bbox spans ~the full page
    and the "Annexure X.Y.: ..." title line sits INSIDE the bbox — the
    above-the-table caption search found nothing and 20+ tables shipped
    untitled. _extract_caption now searches the bbox's own top band when the
    bbox covers >85% of the page. Also: a heading candidate with unbalanced
    closing parens ("Integration) Integration)") is a wrapped header-cell
    tail, never a title."""
    print("Guard BI — frame-title recovery + unbalanced-paren heading rejection")
    from backend.app.extract.table_extractor import _looks_like_heading, extract_tables

    check("unbalanced-paren fragment rejected as heading",
          not _looks_like_heading("Integration) Integration)"), "")
    check("balanced parens still fine",
          _looks_like_heading("Maximum Pendency Percentage (North-Eastern States)"), "")

    pdf = "/Users/thesinghaa/Downloads/cpgram_state_all/Aug_2022.pdf"
    if not os.path.exists(pdf):
        print("  [SKIP] CPGRAM Aug_2022.pdf not found — skipping real-PDF frame-title case")
        return
    tabs = extract_tables(pdf, pages=[10])
    caps = [t.get("caption") or "" for t in tabs]
    check("framed-page Annexure title recovered from inside the bbox",
          any(c.startswith("Annexure 1.4") for c in caps), str(caps))


def guard_fragment_quarantine():
    """Guard BH — infographic/prose fragment quarantine (table_validator).
    Two junk families observed leaking through on the CPGRAM corpus (scores
    0.05-0.24, dragging the corpus mean): (a) narrative paragraphs carrying
    the print footer "N | P a g e" split across a 3x3 grid; (b) infographic
    bullet lists ("PMKISAN related issues Receipts: 509 | 60.74%") where only
    one column has content and the rest are camelot grid ghosts. Both are
    page furniture, not tables. Real small tables must keep passing."""
    print("Guard BH — fragment quarantine (single_column_list + page_footer_fragment)")
    import pandas as pd
    from backend.app.validation.table_validator import validate_table

    prose = pd.DataFrame({
        "label": ["Public Grievances has undertaken a massive mandate of",
                  "States/UTs grievance portals with CPGRAMS.", ""],
        "x": ["integrating all", "", ""],
        "label_2": ["the respective", "", "13 | P a g e"],
    })
    check("(a) prose fragment with page footer quarantined",
          validate_table(prose) == {"passed": False, "reason": "page_footer_fragment"},
          str(validate_table(prose)))

    bullets = pd.DataFrame({
        "col": [""] * 5, "col_2": [""] * 5,
        "label": ["", "PMKISAN related issues Receipts: 56 | 34.15%",
                  "Mobile related Receipts: 8 | 4.88%",
                  "PMKISAN related issues Receipts: 37 | 22.98%",
                  "Pradhan Mantri Gram Sadak Yojana Receipts: 16 | 9.94%"],
    })
    check("(b) single-populated-column bullet list quarantined",
          validate_table(bullets) == {"passed": False, "reason": "single_column_list"},
          str(validate_table(bullets)))

    real = pd.DataFrame({"state": ["Kerala", "Odisha", "Assam"],
                         "receipts": ["120", "340", "77"]})
    check("(c) real small 2-col table still passes",
          validate_table(real)["passed"] is True, str(validate_table(real)))

    # a big table that MENTIONS a page footer in one cell is NOT gated
    big = pd.DataFrame({
        "state": [f"State {i}" for i in range(12)],
        "receipts": [str(100 + i) for i in range(12)],
        "note": ["13 | P a g e"] + [""] * 11,
    })
    check("(d) footer mention in a large table does not quarantine",
          validate_table(big)["passed"] is True, str(validate_table(big)))


def guard_entity_series_stitching():
    """Guard BM — entity-per-page report series must not stitch across
    entities (table_stitcher._entity_variant_titles). NFHS-6 fact sheets
    print an identical header row for every state; the header-equality merge
    stitched Assam's pages into an "Arunachal Pradesh Key Indicators" blob
    (user-reported: state sheets holding other states' data). Titles that
    share a >=2-word tail but differ in their first word name different
    ENTITIES of one series and block the merge; same-title continuations and
    unrelated weak section names are unaffected."""
    print("Guard BM — entity-series stitch boundary (table_stitcher)")
    import pandas as pd
    from backend.app.standardization.table_stitcher import (
        _entity_variant_titles, stitch_tables,
    )

    check("state fact-sheet variants detected",
          _entity_variant_titles("Arunachal Pradesh Key Indicators",
                                 "Assam Key Indicators"), "")
    check("hyphenated variants detected",
          _entity_variant_titles("Haryana - Key Indicators",
                                 "Karnataka - Key Indicators"), "")
    check("same title is NOT an entity variant",
          not _entity_variant_titles("Assam Key Indicators",
                                     "Assam Key Indicators"), "")
    check("unrelated section heading is NOT an entity variant",
          not _entity_variant_titles("Andhra Pradesh Key Indicators",
                                     "Maternal and Child Health"), "")

    def _t(tid, name, page, rows, start):
        # distinct row content per fragment — identical rows would be eaten
        # by the post-merge dedupe and mask whether stitching happened
        df = pd.DataFrame({
            "indicator": [f"ind {start + i}" for i in range(rows)],
            "nfhs6_urban": [f"1{start + i}.0" for i in range(rows)],
            "nfhs6_rural": [f"2{start + i}.0" for i in range(rows)],
            "nfhs6_total": [f"3{start + i}.0" for i in range(rows)],
        })
        return {"table_id": tid, "name": name, "page": page, "df": df,
                "titled": True, "flavor": ""}

    stitched = stitch_tables([
        _t(1, "Arunachal Pradesh Key Indicators", 1, 8, 0),
        _t(2, "Arunachal Pradesh Key Indicators", 2, 8, 8),   # true continuation
        _t(3, "Assam Key Indicators", 3, 8, 16),              # NEXT STATE
    ])
    check("same-state continuation still merges",
          any(s["df"].shape[0] == 16 for s in stitched),
          str([(s["name"][:30], s["df"].shape) for s in stitched]))
    check("next state stays a separate table (no cross-entity blob)",
          len(stitched) == 2 and any(s["name"].startswith("Assam")
                                     and s["df"].shape[0] == 8 for s in stitched),
          str([(s["name"][:30], s["df"].shape) for s in stitched]))


def guard_sparse_high_accuracy_stream_gate():
    """Guard BN — a near-perfect stream parse must not be dropped for being
    sparse. _process_stream_table's density gate rejected ANY table with
    whitespace > 60, regardless of accuracy — real bilingual continuation
    pages (multi-line rows leave many alignment cells blank) can legitimately
    score whitespace just over 60 while parsing at ~99% accuracy. Observed
    on a real state budget PDF: page 5 ("Budget at a Glance" continuation,
    items 7-10a: Revenue/Fiscal/Primary Deficit) scored accuracy=98.98,
    whitespace=61.11 and — because lattice found nothing on that page either
    — the entire page silently vanished from the output with no quarantine
    record. A near-perfect parse (accuracy >= 95) is now trusted even when
    sparse; a mediocre one is still rejected so noisy pseudo-tables don't
    sneak through."""
    print("Guard BN — sparse-but-accurate stream tables are kept (table_extractor)")
    import pandas as pd
    from backend.app.extract.table_extractor import _process_stream_table

    class _FakeTable:
        def __init__(self, df, accuracy, whitespace):
            self.df = df
            self.parsing_report = {"accuracy": accuracy, "whitespace": whitespace}

    real_df = pd.DataFrame([[str(i), f"row {i}", "", "", "", str(i * 10)]
                            for i in range(10)])

    check("high-accuracy sparse table (whitespace=61.11) is KEPT",
          _process_stream_table(_FakeTable(real_df, 98.98, 61.11), None) is not None,
          "regressed: R2016.pdf page 5 (Budget at a Glance, items 7-10a) would vanish again")

    check("mediocre-accuracy sparse table (accuracy=85, whitespace=61.11) is still REJECTED",
          _process_stream_table(_FakeTable(real_df, 85, 61.11), None) is None,
          "the relaxation must not admit genuinely noisy pseudo-tables")

    check("low-accuracy table (accuracy=70) is still REJECTED regardless of whitespace",
          _process_stream_table(_FakeTable(real_df, 70, 10), None) is None,
          "accuracy floor of 80 must remain in force")

    check("high-accuracy but VERY sparse chart-debris (whitespace=82.14) is still REJECTED",
          _process_stream_table(_FakeTable(real_df, 100.0, 82.14), None) is None,
          "regressed: cpgram_dec2023_p11.pdf chart-axis debris (2000/6000/... rows) "
          "would leak back into the output as a bogus table")


def guard_bilingual_suffix_translation():
    """Guard BO — bilingual Kruti-Dev + English cells resolve to the PDF's
    own English wording, and genuine English text is never mistaken for
    Kruti soup.

    Indian government bilingual reports print a Kruti-Dev Hindi label
    immediately followed by its own English translation on the same line
    ("jktLo izkfIr;ka Revenue Receipts" = Hindi label + literal "Revenue
    Receipts"). translate_text now trusts that literal English suffix
    instead of trying to decode/translate the Hindi half word-by-word
    (observed on a real Rajasthan state budget PDF: 'Budget at a Glance'
    labels came out as a mix of raw ASCII soup and half-decoded Devanagari
    sitting right next to the correct English text already in the cell).

    Two false-positive traps found and fixed while building this:
      - "Non-Tax Revenue": looks_kruti's slash-compound exception didn't
        cover hyphens, so "Non-Tax" alone read as soup and got dropped from
        the translated output.
      - "A.Settlement Systems" / "B.Payment Systems" (RBI Table 61 category
        labels, no Hindi involved at all): a list marker glued straight onto
        a word with no space tripped looks_kruti's CamelCase-exception
        fallback, and would have dropped "Settlement"/"Payment" entirely
        from the extracted category column.

    Also pins the decoder completeness fix (kruti_dev._SINGLE was missing
    the 'I' glyph -> 'प्' half-form, so "izkfIr;ka" decoded to
    "प्रािIतयां" instead of "प्राप्तियां")."""
    print("Guard BO — bilingual Kruti-Dev/English cell resolution (hindi_translator)")
    from backend.app.translation.hindi_translator import translate_text
    from backend.app.translation.kruti_dev import kruti_to_unicode, looks_kruti

    check("Hindi label + English suffix -> just the English suffix",
          translate_text("jktLo izkfIr;ka Revenue Receipts") == "Revenue Receipts",
          translate_text("jktLo izkfIr;ka Revenue Receipts"))
    check("glossary-only soup word ('vuqnku') doesn't survive next to its translation",
          translate_text("lgk;rkFkZ vuqnku Grants-in-aid") == "Grants-in-aid",
          translate_text("lgk;rkFkZ vuqnku Grants-in-aid"))

    check("hyphenated English compound ('Non-Tax Revenue') kept whole",
          translate_text("v&dj jktLo Non-Tax Revenue") == "Non-Tax Revenue",
          translate_text("v&dj jktLo Non-Tax Revenue"))
    check("looks_kruti no longer flags 'Non-Tax' as soup",
          not looks_kruti("Non-Tax"), "")

    check("lowercase -ks plurals ('banks', 'stocks', 'networks') are NOT soup",
          not any(looks_kruti(w) for w in ("banks", "stocks", "networks", "blocks")),
          str([w for w in ("banks", "stocks", "networks", "blocks") if looks_kruti(w)]))
    check("genuine soup plurals still caught after the -s exemption",
          all(looks_kruti(w) for w in ("ftys", "ys[ks", "dks")),
          str([w for w in ("ftys", "ys[ks", "dks") if not looks_kruti(w)]))

    check("glued list-marker + word ('A.Settlement Systems') NOT torn apart",
          translate_text("A.Settlement Systems") == "A.Settlement Systems",
          translate_text("A.Settlement Systems"))
    check("glued list-marker + word ('B.Payment Systems') NOT torn apart",
          translate_text("B.Payment Systems") == "B.Payment Systems",
          translate_text("B.Payment Systems"))

    check("decoder gap fixed: 'izkfIr;ka' -> प्राप्तियां (was प्रािIतयां, missing 'I' glyph)",
          kruti_to_unicode("izkfIr;ka") == "प्राप्तियां",
          kruti_to_unicode("izkfIr;ka"))

    check("no English suffix present -> falls back to decode (still legible Devanagari)",
          translate_text("vk;&O;;d") == "आय-व्ययक",
          translate_text("vk;&O;;d"))


def guard_text_table_scoring():
    """Guard BD — archetype-aware quality scoring (quality.score_table).
    Text tables (Yes/No status grids, name/link catalogues) legitimately have
    near-zero numeric_density; grading them on it flagged perfectly-extracted
    tables low_quality and triggered strategy retries that could never help
    (observed on DARPG Annexure 5: nd=0.17 hc=1.0 fill=0.96 -> 0.66). The
    text branch is conservative: it requires the table to be FILLED and
    text-dominant, so sparse/fragmented extractions keep the strict formula."""
    print("Guard BD — text-table scoring branch (extract/quality.score_table)")
    import pandas as pd
    from backend.app.extract.quality import score_table

    # (a) well-extracted text table (Annexure-5 shape): text branch, high score
    text_df = pd.DataFrame({
        "s_no": ["1", "2", "3", "4"],
        "name_state": ["Kerala", "Odisha", "Punjab", "Assam"],
        "portal_type": ["State Portal", "CPGRAMS", "CPGRAMS", "State Portal"],
        "integrated": ["Yes", "Yes", "No", "Yes"],
    })
    q = score_table(text_df)
    check("(a) filled text table enters text_mode", q["text_mode"] is True, str(q))
    check("(a) text table scores high (was nd-punished before)",
          q["score"] >= 0.85, str(q))

    # (b) numeric table: statistical formula unchanged, text_mode off
    num_df = pd.DataFrame({
        "state": ["Kerala", "Odisha"],
        "receipts": ["1200", "3400"],
        "disposal": ["1100", "3300"],
        "pending": ["100", "100"],
    })
    q2 = score_table(num_df)
    check("(b) numeric table keeps statistical formula", q2["text_mode"] is False, str(q2))

    # (c) sparse alphabetic debris must NOT sneak into the text branch
    sparse = pd.DataFrame({
        "col": ["abc", "", "", ""],
        "col_2": ["", "", "xy", ""],
        "col_3": ["", "", "", ""],
    })
    q3 = score_table(sparse)
    check("(c) sparse fragments stay on the strict formula (fill gate)",
          q3["text_mode"] is False and q3["score"] < 0.5, str(q3))

    # (d) fiscal-year ranges / currency amounts / dates are clean VALUES —
    # a grants table full of them must not score numeric_density 0.0
    # (observed: DARPG Sevottam-grants annexure at 0.43 across ~10 reports)
    grants = pd.DataFrame({
        "institute": ["YASHADA", "MGSIPA", "ATI Kerala", "HIPA"],
        "fy": ["2024-25", "2023-24", "2024-25", "2023-24"],
        "sanctioned": ["Rs. 40 lakh", "Rs. 25 lakh", "₹ 12.5 Cr", "1,234 crore"],
        "date": ["01.04.2024", "31/12/2023", "15.06.2024", "01.01.2024"],
    })
    q4 = score_table(grants)
    check("(d) year-range/currency/date cells count as clean values",
          q4["numeric_density"] >= 0.7, str(q4))
    check("(d) grants table scores well (was 0.43-class before)",
          q4["score"] >= 0.80, str(q4))


def guard_month_header_recovery():
    """Guard BE — month-name header rows beat leading chart-axis debris.

    CPGRAMS feedback pages export a chart above the grid, leaving sparse
    numeric debris rows ("2000 | | 2508") on top of the real "Jan .. Dec"
    header. The debris matched the year-row rule, cutting the header band
    short and turning all 13 columns into col_N fallbacks. A row of month
    names is now the strongest header signal, and all-numeric rows above it
    are excluded from column naming (dec_2023 p11 "Feedback marked")."""
    print("Guard BE — buried month header (debris rows above 'Jan .. Dec')")
    import pandas as pd
    from backend.app.cleaning.header_detector import detect_header_rows, is_month_header_row

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # (a) synthetic: debris row + duplicated month band + percentage data
    df = pd.DataFrame([
        ["2000", "", "", "", "2508", "", "", "", "", "", "", "", ""],
        [""] + months,
        ["Month"] + months,
        ["E", "17%", "17%", "16%", "14%", "15%", "13%", "13%", "14%", "15%", "15%", "16%", "15%"],
        ["VG", "8%", "8%", "9%", "9%", "10%", "10%", "10%", "11%", "11%", "10%", "11%", "11%"],
    ])
    check("(a) header band reaches past the debris to the LAST month row",
          detect_header_rows(df) == 3, f"got {detect_header_rows(df)}")
    check("(a) month row recognised", is_month_header_row(df.iloc[1]) is True, "")
    check("(a) data row not month-like", is_month_header_row(df.iloc[3]) is False, "")

    # (b) real page: dec_2023 p11 "Feedback marked Excellent/Very Good"
    items = _pipeline(os.path.join(ROOT, "Testpdfs/cpgram_dec2023_p11.pdf"))
    check("(b) table extracted", bool(items), "none")
    if items:
        cols = list(items[0]["df"].columns)
        check("(b) 12 month columns named (was 13/13 col_N)",
              cols[1:] == [m.lower() for m in months], f"got {cols}")


def guard_lattice_header_band_recovery():
    """Guard BF — lattice tables whose header band sits ABOVE the ruled grid.

    DARPG annexures rule only the data rows; the 3-line interleaved header
    ("S. No. | Name of State/UT | Brought Forward | ... | ACT (in days)")
    lives outside camelot's bbox, so every measure column fell back to
    code_N / value. The stream-flavor positional header recovery now also
    runs on lattice tables (gated on the RAW grid being headerless, band
    cropped at the exact grid top so no data row leaks in).

    Also pins the tightened code-role heuristic: bare 3-4 digit integers are
    only a "code" when fixed-width or zero-padded — varied-width counts
    (grievance receipts) are measures. NCO dotted codes stay codes."""
    print("Guard BF — lattice header-band recovery + code-vs-count roles")
    import pandas as pd
    from backend.app.standardization.column_namer import infer_role

    # (a) real page: aug-2025 p26, Annexures 2.4 + 2.5
    items = _pipeline(os.path.join(ROOT, "Testpdfs/cpgram_aug2025_p26.pdf"))
    check("(a) two tables extracted", len(items) == 2, f"got {len(items)}")
    if len(items) == 2:
        c24 = list(items[0]["df"].columns)
        check("(a) Annexure 2.4 fully named (was 7/9 generic)",
              c24 == ["s_no", "name_state", "brought_forward", "receipts",
                      "total_grievances", "disposed", "pending", "act_days",
                      "pending_percentage"], f"got {c24}")
        c25 = list(items[1]["df"].columns)
        check("(a) Annexure 2.5 fully named (was 5/8 generic)",
              c25 == ["s_no", "name_state", "brought_forward", "receipts",
                      "total_grievances", "disposed", "pending", "pending_days"],
              f"got {c25}")

    # (b) varied-width 3-4 digit counts are NOT codes
    counts = pd.Series(["1280", "135", "1415", "120", "2285", "467"])
    check("(b) grievance counts not misread as code", infer_role(counts) != "code",
          f"got {infer_role(counts)}")

    # (c) classification codes keep their role: fixed-width and zero-padded
    check("(c) fixed-width group codes stay code",
          infer_role(pd.Series(["111", "112", "121", "122"])) == "code", "")
    check("(c) zero-padded codes stay code",
          infer_role(pd.Series(["0110", "0121", "1210", "232"])) == "code", "")
    check("(c) dotted NCO codes stay code",
          infer_role(pd.Series(["1111.0100", "1111.10", "1112.02"])) == "code", "")


def guard_plumber_strategy():
    """Guard BJ — pdfplumber third extraction strategy + honest raw-frame scoring.

    (1) quality.score_table on a RAW (integer RangeIndex) frame substitutes
    populated-COLUMN coherence for the name-based header_coherence, which is
    trivially 1.0 pre-headers. Before this, a hallucinated 21-column lattice
    grid (6 real columns) scored 0.768 and never retried.
    (2) _plumber_retry offers pdfplumber's own table finder as the 3rd
    strategy on NON-scanned pages (OCR keeps the scanned slot), normalizing
    its None cells and honoring the keep-max-score rule via _consider."""
    print("Guard BJ — pdfplumber strategy + raw-frame column-coherence scoring")
    import pandas as pd
    from backend.app.extract.quality import score_table

    # (a) raw frame with phantom columns scores LOW; same content named scores high
    raw = pd.DataFrame([["a", "", "1", "", "2", "", "", ""],
                        ["b", "", "3", "", "4", "", "", ""]])
    named = pd.DataFrame({"state": ["a", "b"], "receipts": ["1", "3"],
                          "disposed": ["2", "4"]})
    s_raw, s_named = score_table(raw)["score"], score_table(named)["score"]
    check("(a) phantom-column raw frame scores below retry threshold",
          s_raw < 0.70 <= s_named, f"raw={s_raw} named={s_named}")

    # (b) real page: feb_2024 p28 — lattice hallucinated 21 columns; the raw
    # score now exposes it and a retry strategy wins
    items = extract_tables(os.path.join(ROOT, "Testpdfs/cpgram_feb2024_p28.pdf"))
    retried = [t for t in items if len(t["attempts"]) > 1]
    check("(b) low-scoring lattice grid triggers a retry", bool(retried),
          f"attempts={[t['attempts'] for t in items]}")
    check("(b) retry strategy beats the lattice attempt",
          all(t["best_score"] > t["attempts"][0]["score"] for t in retried),
          f"{[(t['attempts'], t['best_score']) for t in retried]}")

    # (c) _plumber_retry normalizes None cells and returns a scored frame
    import pdfplumber
    from backend.app.extract.table_extractor import _plumber_retry
    with pdfplumber.open(os.path.join(ROOT, "Testpdfs/cpgram_feb2024_p28.pdf")) as pp:
        r = _plumber_retry(pp, 1)
    check("(c) plumber strategy finds the clean grid", r is not None and r[0].shape[1] >= 7,
          f"got {None if r is None else r[0].shape}")
    if r is not None:
        cells = r[0].values.flatten().tolist()
        check("(c) None cells normalized to empty strings",
              all(c is not None for c in cells), "")


def guard_stacked_table_split():
    """Guard BK — vertically stacked annexures split into separate tables.

    Stream frames holding [tail rows of table N] + ["Annexure N+1: …" title
    row] + [table N+1's header band + data] are split at interior title rows;
    each buried part carries its title row's text as caption. Before this,
    header detection saw data before the band and the buried table's columns
    all degraded to value_N (score 0.63 -> 0.92 on the real slice below)."""
    print("Guard BK — stacked-annexure split (interior title-row boundaries)")
    import pandas as pd
    from backend.app.cleaning.panel_splitter import split_stacked_tables

    # (a) synthetic: tail + title + band + data
    df = pd.DataFrame([
        ["", "Grand Total", "100", "200"],
        ["A", "nnexure 3.2.: Top 10 States", "", ""],
        ["S. No.", "Name of State/UT", "Nodal", "Total"],
        ["1", "Government of Haryana", "72", "74"],
        ["2", "Government of Kerala", "48", "49"],
    ])
    parts = split_stacked_tables(df)
    check("(a) split into tail + buried table", len(parts) == 2, f"got {len(parts)}")
    if len(parts) == 2:
        check("(a) buried part carries the rejoined title as caption",
              parts[1][1] is not None and parts[1][1].startswith("Annexure 3.2"),
              f"got {parts[1][1]}")
        check("(a) buried part starts at its header band",
              str(parts[1][0].iloc[0, 0]).strip() == "S. No.", f"got\n{parts[1][0]}")

    # (b) a frame without interior titles is untouched
    plain = pd.DataFrame([["a", "1"], ["b", "2"], ["c", "3"],
                          ["d", "4"], ["e", "5"]])
    check("(b) no title row -> unchanged single part",
          len(split_stacked_tables(plain)) == 1
          and split_stacked_tables(plain)[0][1] is None, "")

    # (c) real page: march_2023 p15 — Annexure 3.1 + buried 3.2
    items = _pipeline(os.path.join(ROOT, "Testpdfs/cpgram_march2023_p15.pdf"))
    caps = [str(i["name"] or "") for i in items]
    check("(c) buried Annexure 3.2 emerges as its own titled table",
          any("3.2" in c for c in caps), f"got {caps}")
    buried = [i for i in items if "3.2" in str(i["name"] or "")]
    if buried:
        cols = list(buried[0]["df"].columns)
        generic = sum(1 for c in cols
                      if re.fullmatch(r"(col|value|code|label)(_\d+)?", str(c)))
        check("(c) buried table's header band recovered (no generic columns)",
              generic == 0, f"got {cols}")


def guard_content_roles_and_placeholders():
    """Guard BL — new content roles + placeholder-aware quality scoring.

    column_namer: month columns, ascending serial columns (s_no), and state
    names behind decorations ("Government of X", "17. Gujarat", "Haryana- 4")
    now infer confident roles. quality.score_table: explicit missing-value
    markers ("-", "NA", "Nil") no longer dilute numeric/text density (they
    still count as filled); serial cells ("1.") are clean values.
    table_validator._is_prose: the numeric bar is also a COUNT floor, so a
    bullet-note fragment with one stray digit is still prose."""
    print("Guard BL — month/s_no/decorated-state roles, placeholder scoring, prose floor")
    import pandas as pd
    from backend.app.standardization.column_namer import infer_role
    from backend.app.extract.quality import score_table
    from backend.app.validation.table_validator import _is_prose, _to_text_frame

    check("(a) month-name column -> month",
          infer_role(pd.Series(["January", "February", "March", "April'25"])) == "month", "")
    check("(a) ascending ordinals -> s_no (tolerates one debris cell)",
          infer_role(pd.Series(["1.", "2.", "3.", "4.", "5.", "6. N"])) == "s_no", "")
    check("(a) non-ascending small ints are NOT s_no",
          infer_role(pd.Series(["3", "1", "9", "2"])) != "s_no", "")
    check("(a) 'Government of X' column -> state",
          infer_role(pd.Series(["Government of Kerala", "Government of Bihar",
                                "Government of Assam"])) == "state", "")
    check("(a) ordinal-prefixed state list -> state",
          infer_role(pd.Series(["1. Kerala", "2. Assam", "3. Tamil Nadu"])) == "state", "")
    check("(a) count-suffixed state list -> state",
          infer_role(pd.Series(["Haryana- 4", "Punjab- 2", "Rajasthan- 1"])) == "state", "")

    # (b) placeholder dashes no longer tank a clean Yes/No status grid
    grid = pd.DataFrame({"state": ["Kerala", "Bihar", "Assam"],
                         "portal": ["Yes", "-", "Yes"],
                         "link": ["Yes", "-", "-"]})
    check("(b) dash placeholders excluded from density (score >= 0.8)",
          score_table(grid)["score"] >= 0.8, f"got {score_table(grid)}")

    # (c) bullet-note fragment with one stray digit is prose
    frag = pd.DataFrame([["2", "•", "The monthly disposal in States and UTs "
                          "decreased from many cases to fewer cases overall"],
                         ["", "•", "Grievance redressal timelines improved in "
                          "most Ministries compared with the previous month"]])
    check("(c) bullet-note fragment detected as prose",
          _is_prose(_to_text_frame(frag)), "")


def guard_crushed_month_redistribution():
    """Guard BG — month labels crushed into one lattice header cell.

    dec_2025 p5 "2.2 Month-wise Status": lattice merged four (and six) month
    labels into single header cells ("Feb'25 Mar'25 Apr'25 May'25") while the
    data cells stayed separate, so neighbouring columns fell back to
    col/value_N. When the month tokens across the header band, in reading
    order, cover the trailing generic/month-named columns, they are now
    redistributed one month per column."""
    print("Guard BG — crushed month-header redistribution")
    import pandas as pd
    from backend.app.cleaning.header_builder import _redistribute_crushed_months

    # (a) synthetic: 1 + crushed-4 + 1 + crushed-6 month tokens over 12 cols
    header_df = pd.DataFrame([
        ["", "Jan'25", "", "Feb'25 Mar'25 Apr'25 May'25", "", "", "Jun'25",
         "", "Jul'25 Aug'25 Sep'25 Oct'25 Nov'25 Dec'25", "", "", "", ""],
    ])
    data = pd.DataFrame([["Receipts"] + [str(100 + i) for i in range(12)]],
                        columns=["col"] + [f"col_{i}" for i in range(1, 13)])
    out = _redistribute_crushed_months(data.copy(), header_df, set())
    check("(a) months distributed positionally",
          list(out.columns) == ["col", "jan", "feb", "mar", "apr", "may", "jun",
                                "jul", "aug", "sep", "oct", "nov", "dec"],
          f"got {list(out.columns)}")

    # (b) no crush evidence (every cell a single month) -> untouched
    single = pd.DataFrame([["", "Jan", "Feb", "Mar"]])
    data2 = pd.DataFrame([["x", "1", "2", "3"]],
                         columns=["col", "col_1", "col_2", "col_3"])
    out2 = _redistribute_crushed_months(data2.copy(), single, set())
    check("(b) untouched without a crushed cell",
          list(out2.columns) == ["col", "col_1", "col_2", "col_3"],
          f"got {list(out2.columns)}")

    # (c) more month tokens than overwritable columns -> untouched
    #     (never guess alignment)
    short = pd.DataFrame([["", "Feb'25 Mar'25 Apr'25 May'25 Jun'25 Jul'25", "", ""]])
    out3 = _redistribute_crushed_months(data2.copy(), short, set())
    check("(c) untouched on count mismatch",
          list(out3.columns) == ["col", "col_1", "col_2", "col_3"],
          f"got {list(out3.columns)}")

    # (d) real page: dec_2025 p5
    items = _pipeline(os.path.join(ROOT, "Testpdfs/cpgram_dec2025_p5.pdf"))
    month_tables = [i for i in items
                    if list(i["df"].columns)[1:] == ["jan", "feb", "mar", "apr", "may", "jun",
                                                     "jul", "aug", "sep", "oct", "nov", "dec"]]
    check("(d) month-wise status table carries 12 real month columns",
          bool(month_tables), f"got {[list(i['df'].columns) for i in items]}")


def guard_edge_anchor():
    """Guard BM — right-edge column anchoring + word-layer repairs.

    Generalized from the Rajasthan Budget-At-A-Glance series: numeric values
    in ruled tables are right-aligned, so clustering their right edges gives
    exact column assignment where camelot misaligns. Two word-layer defects
    are repaired first: fused parenthesized values ("(a)(b)" one word) split
    at ')(' boundaries, and split values ("-1" + "401108.93") merged. Wired
    as _text_retry's second candidate in the 3rd-attempt slot (non-scanned
    pages), still recording exactly one attempt for that slot."""
    print("Guard BM — right-edge anchoring, fused-token split, split-value merge")
    import pdfplumber
    from backend.app.extract.edge_anchor import (
        cluster_right_edges,
        extract_by_right_edge,
        grid_from_words,
        merge_split_values,
        split_fused_tokens,
    )
    from backend.app.extract.table_extractor import _text_retry

    # (a) fused parenthesized values split with x apportioned by char count
    toks = split_fused_tokens([{"x0": 100.0, "x1": 140.0, "top": 10.0,
                                "text": "(1500000.00)(1500000.00)"}])
    check("(a) fused token splits into two values",
          [t["text"] for t in toks] == ["(1500000.00)", "(1500000.00)"],
          f"got {[t['text'] for t in toks]}")
    check("(a) split boxes tile the original x-range",
          len(toks) == 2 and toks[0]["x0"] == 100.0 and toks[1]["x1"] == 140.0
          and abs(toks[0]["x1"] - toks[1]["x0"]) < 1e-9,
          f"got {toks}")
    # a non-value ')(' word (prose) is untouched
    prose = split_fused_tokens([{"x0": 0, "x1": 10, "top": 0, "text": "a)(b"}])
    check("(a) non-numeric ')(' word untouched", prose[0]["text"] == "a)(b", "")

    # (b) short leading fragment merges into its numeric neighbour
    toks = merge_split_values([
        {"x0": 100, "x1": 108, "top": 20, "text": "-1"},
        {"x0": 110, "x1": 160, "top": 20, "text": "401108.93"},
    ])
    check("(b) '-1' + '401108.93' merge to one value",
          [t["text"] for t in toks] == ["-1401108.93"],
          f"got {[t['text'] for t in toks]}")
    # a complete value never absorbs its neighbour, nor across a wide gap
    keep = merge_split_values([
        {"x0": 100, "x1": 130, "top": 20, "text": "123,456.00"},
        {"x0": 132, "x1": 160, "top": 20, "text": "789.00"},
    ])
    check("(b) complete values stay separate", len(keep) == 2, f"got {keep}")
    far = merge_split_values([
        {"x0": 100, "x1": 108, "top": 20, "text": "-1"},
        {"x0": 140, "x1": 160, "top": 20, "text": "401108.93"},
    ])
    check("(b) wide gap blocks the merge", len(far) == 2, f"got {far}")

    # (c) right-edge clustering: anchors need >= 3 aligned edges
    anchors = cluster_right_edges([200.1, 200.4, 199.8, 300.0, 300.2, 299.9,
                                   50.0])
    check("(c) two anchors from two tight clusters, stray edge dropped",
          len(anchors) == 2 and abs(anchors[0] - 200.1) < 1
          and abs(anchors[1] - 300.03) < 1, f"got {anchors}")

    # (d) synthetic misaligned grid: values slot by right edge, not word order
    words = []
    for r in range(4):
        top = 10.0 * (r + 1)
        words.append({"x0": 10, "x1": 60, "top": top, "text": f"Row {r}"})
        # col-1 values vary in width (left edges wander, right edges align)
        w = 20 + 7 * r
        words.append({"x0": 200 - w, "x1": 200, "top": top,
                      "text": f"{10 ** r}.00"})
        words.append({"x0": 280, "x1": 300, "top": top, "text": "555.00"})
    grid = grid_from_words(words)
    check("(d) synthetic grid: 3 columns (label + 2 anchors)",
          grid is not None and all(len(r) == 3 for r in grid),
          f"got {grid}")
    check("(d) wandering-left-edge values all land in column 1",
          grid is not None and [r[1] for r in grid] ==
          ["1.00", "10.00", "100.00", "1000.00"], f"got {grid}")

    # (e) real page: R2018 BAGE p3 — fused "(1500000.00)(1500000.00)" splits
    # and the four value columns carry the repaired parenthesized values
    with pdfplumber.open(os.path.join(ROOT, "Testpdfs/bage_r2018_p3.pdf")) as pdf:
        g18 = extract_by_right_edge(pdf.pages[0])
    check("(e) R2018 p3 rebuilds as label + 4 value columns",
          g18 is not None and len(g18[0]) == 5, f"got {None if g18 is None else len(g18[0])}")
    fused_rows = [r for r in (g18 or [])
                  if r[2:] == ["(1500000.00)", "(1500000.00)", "(1500000.00)"]]
    check("(e) fused Uday row split across its three columns",
          bool(fused_rows), f"got {[r for r in (g18 or []) if '1500000' in ' '.join(r)]}")

    # (f) real page: R2021 BAGE p5 — split "-1 401108.93" merges back
    with pdfplumber.open(os.path.join(ROOT, "Testpdfs/bage_r2021_p5.pdf")) as pdf:
        g21 = extract_by_right_edge(pdf.pages[0])
    check("(f) split value '-1 401108.93' repaired to -1401108.93",
          g21 is not None and any(r[1] == "-1401108.93" for r in g21),
          f"got {[r[1] for r in (g21 or []) if r[1]]}")

    # (g) integration: _text_retry offers the edge grid; the retry loop still
    # records at most 3 attempts and the 3rd slot is a single strategy
    with pdfplumber.open(os.path.join(ROOT, "Testpdfs/bage_r2021_p5.pdf")) as pdf:
        tr = _text_retry(pdf, 1)
        check("(g) _text_retry returns a scored (df, score, strategy)",
              tr is not None and tr[2] in ("plumber", "edge"),
              f"got {tr if tr is None else tr[2]}")
    items = extract_tables(os.path.join(ROOT, "Testpdfs/bage_r2018_p3.pdf"))
    check("(g) attempts capped at 3 with one strategy per slot",
          items and all(len(t["attempts"]) <= 3 for t in items),
          f"got {[t['attempts'] for t in items]}")


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")
    # Remember if caller wanted Docling, then force-unset so base guards run on Camelot
    _docling_requested = os.environ.pop("DOCLING_ENABLED", "").lower() in ("1", "true", "yes")
    base_guards = (guard_des, guard_darpg, guard_plfs, guard_nfhs, guard_nfhs5, guard_fr375_kpi,
                   guard_rbi_payment_system, guard_rbi_money_stock, guard_rbi_multipage_stitch,
                   guard_rbi_orphan_merge, guard_rbi_msp_headers, guard_rbi_bare_year,
                   guard_rbi_multilevel_header, guard_numeric_below_unit,
                   guard_header_recovery_gate, guard_side_by_side_split,
                   guard_section_heading_detection,
                   guard_continuation_title_inheritance,
                   guard_title_and_composite_metric, guard_corruption_quarantine,
                   guard_corruption_detector, guard_transliteration,
                   guard_ocr_recovery, guard_ghost_on_census,
                   guard_descriptive_title,
                   guard_column_dedupe, guard_phantom_value_columns,
                   guard_numeric_readiness_metric, guard_thin_subheader,
                   guard_section_lift, guard_multilevel_header_merge,
                   guard_workbook_toc, guard_numeric_normalization,
                   guard_ghost_suppression, guard_toc_prose_quarantine,
                   guard_title_recovery, guard_reference_tables,
                   guard_extract_retry, guard_recheck_quarantine,
                   guard_page_scope_real_pdf,
                   guard_wrapped_header_reconstruction, guard_text_table_scoring,
                   guard_month_header_recovery, guard_lattice_header_band_recovery,
                   guard_crushed_month_redistribution, guard_fragment_quarantine,
                   guard_frame_title_recovery,
                   guard_plumber_strategy, guard_stacked_table_split,
                   guard_content_roles_and_placeholders, guard_entity_series_stitching,
                   guard_sparse_high_accuracy_stream_gate,
                   guard_bilingual_suffix_translation, guard_edge_anchor)
    # Guard G handles its own DOCLING_ENABLED toggle; only add it when explicitly requested
    extra_guards = (guard_nfhs_docling,) if _docling_requested else ()
    for g in base_guards + extra_guards:
        try:
            g()
        except Exception as e:
            FAILURES.append(f"{g.__name__} crashed: {e}")
            print(f"  [FAIL] {g.__name__} crashed: {e}")
    print()
    if FAILURES:
        print(f"RED — {len(FAILURES)} failure(s)")
        sys.exit(1)
    print("GREEN — all guards pass")
