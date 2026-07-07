"""Outputs for the batch panel builder: a master workbook (one sheet per panel
+ a Schema Changes sheet + a TOC) and a standalone markdown changelog.

Reuses the single-PDF exporter's styling and helpers (`_sheet_name`,
`_maybe_number`, the header/title fonts) so the two workbooks look identical;
the panel workbook only differs in structure (period-coverage TOC + a dedicated
schema-evolution sheet)."""
import io
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.app.export.excel_exporter import (
    HEADER_FILL,
    HEADER_FONT,
    LINK_FONT,
    TITLE_FONT,
    _maybe_number,
    _sheet_name,
)

_KIND_ORDER = ["added", "dropped", "renamed", "combined", "split"]


def _write_panel_sheet(wb, used, sig, panel, label):
    sheet = _sheet_name(label, sig, used)
    ws = wb.create_sheet(sheet)
    n_cols = max(panel.shape[1], 1)
    ws.cell(row=1, column=1, value=str(label)).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for c, col in enumerate(panel.columns, start=1):
        cell = ws.cell(row=3, column=c, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(panel.itertuples(index=False), start=4):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=_maybe_number(value))
    for c, col in enumerate(panel.columns, start=1):
        width = max(len(str(col)),
                    *(len(str(v)) for v in panel.iloc[:, c - 1].head(50)), 8)
        ws.column_dimensions[get_column_letter(c)].width = min(width + 2, 45)
    ws.freeze_panes = "A4"
    return sheet


def build_panel_workbook(panels):
    """panels: list of {signature, label, panel(df), diff} in display order.
    Returns BytesIO of the master .xlsx.

    Multi-period panels (n_periods >= 2) get individual sheets + appear in the
    main Panels TOC. Single-period tables go to a 'Single Period' reference
    sheet so they don't clutter the analyst view."""
    wb = Workbook()
    toc = wb.active
    toc.title = "Panels"
    used = set()
    toc_rows = []
    single_rows = []

    for p in panels:
        if p["panel"] is None or p["panel"].shape[0] == 0:
            continue
        d = p["diff"]
        n_periods = len(d["periods"])
        if n_periods >= 2:
            sheet = _write_panel_sheet(wb, used, p["signature"], p["panel"], p["label"])
            toc_rows.append((p["label"], ", ".join(d["periods"]),
                             p["panel"].shape[0], p["panel"].shape[1],
                             len(d["changes"]), sheet))
        else:
            period = d["periods"][0] if d["periods"] else "?"
            single_rows.append((p["label"], period,
                                 p["panel"].shape[0], p["panel"].shape[1]))

    # --- Panels TOC ---
    headers = ["Panel", "Periods", "Rows", "Cols", "Schema changes", "Sheet"]
    for c, h in enumerate(headers, 1):
        cell = toc.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, (label, periods, nrows, ncols, nchg, sheet) in enumerate(toc_rows, start=2):
        link = toc.cell(row=r, column=1, value=str(label))
        link.hyperlink = f"#'{sheet}'!A1"
        link.font = LINK_FONT
        toc.cell(row=r, column=2, value=periods)
        toc.cell(row=r, column=3, value=nrows)
        toc.cell(row=r, column=4, value=ncols)
        toc.cell(row=r, column=5, value=nchg)
        toc.cell(row=r, column=6, value=sheet)
    for col, w in {"A": 60, "B": 30, "C": 8, "D": 7, "E": 15, "F": 28}.items():
        toc.column_dimensions[col].width = w
    toc.freeze_panes = "A2"
    if not toc_rows:
        toc.cell(row=2, column=1, value="No multi-period panels assembled.")

    # --- Schema Changes sheet (multi-period panels only) ---
    sc = wb.create_sheet("Schema Changes")
    sc_headers = ["Panel", "Period", "Change", "Variables", "Confidence", "Note"]
    for c, h in enumerate(sc_headers, 1):
        cell = sc.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    r = 2
    any_change = False
    for p in panels:
        if len(p["diff"]["periods"]) < 2:
            continue
        for chg in p["diff"]["changes"]:
            any_change = True
            sc.cell(row=r, column=1, value=str(p["label"]))
            sc.cell(row=r, column=2, value=str(chg["period"]))
            sc.cell(row=r, column=3, value=chg["kind"])
            sc.cell(row=r, column=4, value=" ".join(str(v) for v in chg["variables"]))
            sc.cell(row=r, column=5, value=chg["confidence"])
            sc.cell(row=r, column=6, value=chg["note"])
            r += 1
    if not any_change:
        sc.cell(row=2, column=1, value="No schema changes detected across periods.")
    for col, w in {"A": 50, "B": 12, "C": 12, "D": 50, "E": 12, "F": 60}.items():
        sc.column_dimensions[col].width = w
    sc.freeze_panes = "A2"

    # --- Single Period reference sheet ---
    if single_rows:
        sp = wb.create_sheet("Single Period")
        sp_headers = ["Table", "Period", "Rows", "Cols"]
        for c, h in enumerate(sp_headers, 1):
            cell = sp.cell(row=1, column=c, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        sp.cell(row=1, column=5,
                value="These tables appear in only one period — not stacked into panels.").font = HEADER_FONT
        for r, (label, period, nrows, ncols) in enumerate(single_rows, start=2):
            sp.cell(row=r, column=1, value=str(label))
            sp.cell(row=r, column=2, value=period)
            sp.cell(row=r, column=3, value=nrows)
            sp.cell(row=r, column=4, value=ncols)
        for col, w in {"A": 60, "B": 12, "C": 8, "D": 7}.items():
            sp.column_dimensions[col].width = w
        sp.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def write_changelog_md(panels, path):
    """Human-readable schema changelog, grouped by panel then period — suitable
    for a report's methodology appendix."""
    lines = ["# Schema changelog", ""]
    lines.append("Documents how each panel's variables changed across reporting "
                 "periods. Stacked data is aligned to a canonical variable set; "
                 "every detected change is listed with a confidence level. "
                 "`combined` / `split` are heuristic flags to verify, not "
                 "applied transformations.")
    lines.append("")
    for p in panels:
        d = p["diff"]
        lines.append(f"## {p['label']}")
        lines.append("")
        lines.append(f"- Periods: {', '.join(d['periods'])}")
        lines.append(f"- Canonical variables ({len(d['canonical'])}): "
                     + ", ".join(d["canonical"]))
        lines.append("")
        if not d["changes"]:
            lines.append("_No schema changes detected._")
            lines.append("")
            continue
        for kind in _KIND_ORDER:
            rows = [c for c in d["changes"] if c["kind"] == kind]
            if not rows:
                continue
            lines.append(f"### {kind.capitalize()}")
            for c in rows:
                vs = " ".join(str(v) for v in c["variables"])
                lines.append(f"- **{c['period']}** [{c['confidence']}]: {vs} — {c['note']}")
            lines.append("")
    with open(path, "w") as f:
        f.write("\n".join(lines))
    return path
